from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import (
    RehabilitationFacility,
    RehabilitationFacilityContact,
    RehabilitationFacilityLocation,
    RehabilitationUnresolvedField,
    ScrapingBlueprint,
    ScrapingBlueprintStatus,
    ScrapingExecution,
    ScrapingExecutionStatus,
    ScrapingMission,
    ScrapingRun,
    ScrapingRunStatus,
)
from app.services.scraping.execution_export_service import SHEET_ORDER, execution_export_service
from conftest import create_model_set, valid_blueprint


async def create_completed_real_execution(db: AsyncSession, auth) -> ScrapingExecution:
    model_set = await create_model_set(db, auth, slug="export-honesty")
    mission = ScrapingMission(
        org_id=auth.org_id,
        created_by=auth.user.id,
        model_set_id=model_set.slug,
        title="France real rehab export",
        original_prompt="Find real facilities in France",
        country_code="FR",
        country_name="France",
    )
    db.add(mission)
    await db.flush()
    blueprint = ScrapingBlueprint(
        mission_id=mission.id,
        version=1,
        status=ScrapingBlueprintStatus.APPROVED,
        blueprint_json=valid_blueprint(),
        model_set_id=model_set.slug,
        judge_model_id="gpt-4.1",
    )
    db.add(blueprint)
    await db.flush()
    mission.active_blueprint_id = blueprint.id
    run = ScrapingRun(
        organization_id=auth.org_id,
        mission_id=mission.id,
        blueprint_id=blueprint.id,
        model_set_id=model_set.slug,
        status=ScrapingRunStatus.PLANNED,
    )
    db.add(run)
    await db.flush()
    execution = ScrapingExecution(
        organization_id=auth.org_id,
        mission_id=mission.id,
        blueprint_id=blueprint.id,
        team_plan_id=run.id,
        execution_type="initial_full_country",
        mode="real",
        status=ScrapingExecutionStatus.COMPLETED,
        country_code="FR",
        country_name="France",
    )
    db.add(execution)
    await db.flush()
    return execution


@pytest.mark.asyncio
async def test_real_export_uses_classification_sheets_and_honest_copy(
    db: AsyncSession, auth
):
    execution = await create_completed_real_execution(db, auth)

    verified = RehabilitationFacility(
        execution_id=execution.id,
        organization_id=auth.org_id,
        stable_key="verified-centre",
        canonical_name="Centre Verifie",
        original_language_name=None,
        description="Verified public facility",
        facility_type="residential",
        organization_type="private",
        operational_status="active",
        country_code="FR",
        country_name="France",
        primary_region="Ile-de-France",
        primary_city="Paris",
        primary_address="10 Rue Verifie, Paris",
        primary_website="https://verified.example.fr",
        verification_status="verified",
        confidence_score=0.98,
        duplicate_status="unique",
        human_review_status="not_required",
        country_containment_status="inside_target_country",
        country_containment_reason="Matched France address and phone.",
        publication_class="verified",
        is_mock=False,
    )
    review = RehabilitationFacility(
        execution_id=execution.id,
        organization_id=auth.org_id,
        stable_key="review-centre",
        canonical_name="Centre Review",
        original_language_name=None,
        description="Needs more evidence",
        facility_type="outpatient",
        organization_type="private",
        operational_status="active",
        country_code="FR",
        country_name="France",
        primary_region="Ile-de-France",
        primary_city="Paris",
        primary_address="20 Rue Review, Paris",
        primary_website="https://review.example.fr",
        verification_status="unverified",
        confidence_score=0.62,
        duplicate_status="unique",
        human_review_status="required",
        country_containment_status="inside_target_country",
        country_containment_reason="Address appears local but phone is missing.",
        publication_class="review_required",
        is_mock=False,
    )
    excluded = RehabilitationFacility(
        execution_id=execution.id,
        organization_id=auth.org_id,
        stable_key="excluded-centre",
        canonical_name="Centre Excluded",
        original_language_name=None,
        description="Wrong-country result",
        facility_type="residential",
        organization_type="private",
        operational_status="active",
        country_code="DE",
        country_name="Germany",
        primary_region="Berlin",
        primary_city="Berlin",
        primary_address="30 Berliner Platz, Berlin",
        primary_website="https://excluded.example.de",
        verification_status="unverified",
        confidence_score=0.40,
        duplicate_status="unique",
        human_review_status="required",
        country_containment_status="outside_target_country",
        country_containment_reason="Evidence points to Germany instead of France.",
        publication_class="excluded",
        is_mock=False,
    )
    db.add_all([verified, review, excluded])
    await db.flush()

    verified_location = RehabilitationFacilityLocation(
        facility_id=verified.id,
        location_type="campus",
        location_name="Paris Campus",
        country_code="FR",
        country_name="France",
        region="Ile-de-France",
        city="Paris",
        full_address="10 Rue Verifie, Paris",
        is_primary=True,
        verification_status="verified",
        confidence_score=0.98,
        country_containment_status="inside_target_country",
        country_containment_reason="Local address and phone.",
        location_completeness_status="complete",
        location_gap_reason=None,
        is_mock=False,
    )
    review_location = RehabilitationFacilityLocation(
        facility_id=review.id,
        location_type="campus",
        location_name="Review Campus",
        country_code="FR",
        country_name="France",
        region="Ile-de-France",
        city="Paris",
        full_address="20 Rue Review, Paris",
        is_primary=True,
        verification_status="unverified",
        confidence_score=0.62,
        country_containment_status="inside_target_country",
        country_containment_reason="No conflicting country evidence.",
        location_completeness_status="missing_phone",
        location_gap_reason="phone_missing",
        is_mock=False,
    )
    excluded_location = RehabilitationFacilityLocation(
        facility_id=excluded.id,
        location_type="campus",
        location_name="Berlin Campus",
        country_code="DE",
        country_name="Germany",
        region="Berlin",
        city="Berlin",
        full_address="30 Berliner Platz, Berlin",
        is_primary=True,
        verification_status="unverified",
        confidence_score=0.40,
        country_containment_status="outside_target_country",
        country_containment_reason="Outside France.",
        location_completeness_status="complete",
        location_gap_reason=None,
        is_mock=False,
    )
    db.add_all([verified_location, review_location, excluded_location])
    await db.flush()

    db.add_all(
        [
            RehabilitationFacilityContact(
                facility_id=verified.id,
                location_id=verified_location.id,
                contact_type="phone",
                label="Main line",
                value="+33 1 23 45 67 89",
                normalized_value="+33123456789",
                is_primary=True,
                verification_status="verified",
                confidence_score=0.99,
                contact_discovery_status="verified_direct",
                is_mock=False,
            ),
            RehabilitationFacilityContact(
                facility_id=review.id,
                location_id=review_location.id,
                contact_type="email",
                label="Admissions",
                value="hello@review.example.fr",
                normalized_value="hello@review.example.fr",
                is_primary=True,
                verification_status="unverified",
                confidence_score=0.60,
                contact_discovery_status="found_unverified",
                is_mock=False,
            ),
            RehabilitationFacilityContact(
                facility_id=excluded.id,
                location_id=excluded_location.id,
                contact_type="phone",
                label="Main line",
                value="+49 30 123456",
                normalized_value="+4930123456",
                is_primary=True,
                verification_status="unverified",
                confidence_score=0.45,
                contact_discovery_status="found_unverified",
                is_mock=False,
            ),
        ]
    )
    db.add(
        RehabilitationUnresolvedField(
            facility_id=review.id,
            field_path="contacts.phone",
            unresolved_status="conflicting",
            reason="Directory and website disagree on the phone number.",
            recommended_follow_up="Review an official listing.",
            source_id=None,
            is_mock=False,
        )
    )
    await db.commit()

    payload, filename = await execution_export_service.build_workbook(db, auth, execution.id)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == SHEET_ORDER
    assert filename.startswith("scraping-execution-france-")

    facilities_sheet = workbook["Facilities"]
    assert facilities_sheet["A1"].value == "Facility"
    assert facilities_sheet["B1"].value == "Contact"
    assert facilities_sheet["C1"].value == "Website"
    rows = list(facilities_sheet.iter_rows(min_row=2, values_only=True))
    by_name = {row[0]: row for row in rows}
    assert "Centre Verifie" in by_name
    assert "Centre Review" in by_name
    assert "Centre Excluded" not in by_name
    assert by_name["Centre Verifie"][1] == "'+33 1 23 45 67 89"
    assert "mock" not in filename.lower()
    assert "fiction" not in filename.lower()
