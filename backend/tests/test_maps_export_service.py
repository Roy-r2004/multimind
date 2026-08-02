"""Tests for the Maps Census Excel workbook export."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.dependencies import AuthContext
from app.db.models import MapsCensusCell, MapsCensusCellStatus, MapsCensusRun, MapsCensusStatus, MapsPlace
from app.services.scraping.maps_export_service import (
    ELIGIBLE_CENTERS_SHEET,
    EXPORT_HEADERS,
    MIME_XLSX,
    maps_export_service,
)


async def _create_run(db, auth: AuthContext) -> MapsCensusRun:
    run = MapsCensusRun(
        organization_id=auth.org_id,
        created_by=auth.user.id,
        country_code="DZ",
        country_name="Algeria",
        status=MapsCensusStatus.COMPLETED,
    )
    db.add(run)
    await db.flush()
    await db.commit()
    return run


def _workbook(content: bytes):
    return load_workbook(BytesIO(content))


def _sheet_names(sheet):
    return [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]


def _eligible_place(run: MapsCensusRun, **overrides) -> MapsPlace:
    payload = {
        "run_id": run.id,
        "google_place_id": "eligible",
        "raw_name": "Eligible Rehab",
        "canonical_name": "Eligible Rehab",
        "is_relevant": True,
        "confidence_score": 0.95,
        "formatted_address": "1 Rehab Lane, Algiers",
        "client_eligibility": "eligible",
        "lifecycle_status": "confirmed_eligible",
        "keep_drop_decision": "keep",
        "keep_drop_reason": "private rehab confirmed",
        "operator_name": "Eligible Org",
        "operator_type": "association",
        "ownership_status": "confirmed_non_government",
        "funding_type": "public",
        "facility_type": "residential_addiction_rehab",
        "care_setting": "residential",
        "organization_scope": "facility",
        "addiction_focus_confirmed": True,
    }
    payload.update(overrides)
    return MapsPlace(**payload)


@pytest.mark.asyncio
async def test_export_xlsx_single_keep_sheet_only(db, auth):
    run = await _create_run(db, auth)
    db.add(_eligible_place(run))
    db.add(
        MapsPlace(
            run_id=run.id,
            google_place_id="dropped",
            raw_name="Dropped Clinic",
            canonical_name="Dropped Clinic",
            is_relevant=False,
            client_eligibility="excluded",
            lifecycle_status="unrelated",
            keep_drop_decision="drop",
        )
    )
    await db.commit()

    content, filename = await maps_export_service.build_workbook(db, auth, run.id)
    workbook = _workbook(content)

    assert filename == "dz-maps-census-export.xlsx"
    assert workbook.sheetnames == [ELIGIBLE_CENTERS_SHEET]
    assert _sheet_names(workbook[ELIGIBLE_CENTERS_SHEET]) == ["Eligible Rehab"]


@pytest.mark.asyncio
async def test_export_xlsx_eligible_headers_match_spec(db, auth):
    run = await _create_run(db, auth)
    db.add(_eligible_place(run))
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]

    assert [cell.value for cell in sheet[1]] == EXPORT_HEADERS


@pytest.mark.asyncio
async def test_export_xlsx_uses_placeholders_for_missing_values(db, auth):
    run = await _create_run(db, auth)
    db.add(
        _eligible_place(
            run,
            google_place_id="bare",
            canonical_name="Bare Rehab",
            raw_name="Bare Rehab",
            operator_name=None,
            official_website=None,
            international_phone_number=None,
            addictions_treated=None,
            languages_spoken=None,
            verification_source_url=None,
            discovery_sources=None,
            medical_detox=None,
            residential_accommodation=None,
            classification_confidence=None,
        )
    )
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]
    values = [cell.value for cell in sheet[2]]

    assert values.count("Not Specified") >= 4
    assert "Contact for pricing" in values


@pytest.mark.asyncio
async def test_export_xlsx_preserves_arabic_names(db, auth):
    run = await _create_run(db, auth)
    db.add(
        _eligible_place(
            run,
            google_place_id="ar",
            raw_name="مركز بوشاوي للادمان",
            canonical_name="مركز بوشاوي للادمان",
            formatted_address="Chéraga, Algiers",
        )
    )
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]

    assert "مركز بوشاوي للادمان" in _sheet_names(sheet)


@pytest.mark.asyncio
async def test_export_xlsx_website_is_hyperlinked_and_phone_is_text(db, auth):
    run = await _create_run(db, auth)
    db.add(
        _eligible_place(
            run,
            google_place_id="linky",
            official_website="https://linked.example/",
            international_phone_number="+213 555 00 11 22",
            verification_source_url="https://linked.example/evidence",
        )
    )
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]
    header = [cell.value for cell in sheet[1]]
    website_col = header.index("Website") + 1
    phone_col = header.index("Phone Number") + 1

    website_cell = sheet.cell(row=2, column=website_col)
    phone_cell = sheet.cell(row=2, column=phone_col)

    assert website_cell.hyperlink is not None
    assert website_cell.hyperlink.target == "https://linked.example/"
    assert isinstance(phone_cell.value, str)
    assert phone_cell.value.startswith("+213")
    assert phone_cell.number_format == "@"


@pytest.mark.asyncio
async def test_export_xlsx_has_frozen_header_and_autofilter(db, auth):
    run = await _create_run(db, auth)
    db.add(_eligible_place(run))
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


@pytest.mark.asyncio
async def test_export_xlsx_alignment_is_readable(db, auth):
    run = await _create_run(db, auth)
    db.add(
        _eligible_place(
            run,
            google_place_id="align-1",
            raw_name="Alignment Rehab",
            canonical_name="Alignment Rehab",
            formatted_address="1 Long Street Name, Algiers",
            official_website="https://align.example/",
            international_phone_number="+213 555 00 00",
        )
    )
    await db.commit()

    content, _ = await maps_export_service.build_workbook(db, auth, run.id)
    sheet = _workbook(content)[ELIGIBLE_CENTERS_SHEET]
    header = [cell.value for cell in sheet[1]]

    header_cell = sheet.cell(row=1, column=1)
    assert header_cell.alignment.horizontal == "center"
    assert header_cell.alignment.vertical == "center"
    assert header_cell.alignment.wrap_text is True
    assert (sheet.row_dimensions[1].height or 0) >= 24

    name_col = header.index("Facility Name") + 1
    name_cell = sheet.cell(row=2, column=name_col)
    assert name_cell.alignment.horizontal == "left"
    assert name_cell.alignment.vertical == "center"
    assert name_cell.alignment.wrap_text is True

    phone_col = header.index("Phone Number") + 1
    phone_cell = sheet.cell(row=2, column=phone_col)
    assert phone_cell.alignment.horizontal == "center"
    assert phone_cell.alignment.vertical == "center"


@pytest.mark.asyncio
async def test_export_xlsx_returns_mime_and_rejects_other_orgs(db, auth):
    run = await _create_run(db, auth)
    db.add(_eligible_place(run))
    await db.commit()

    assert MIME_XLSX.endswith("spreadsheetml.sheet")

    from app.core.exceptions import NotFoundError

    other = AuthContext(user=auth.user, org_id="different-org", role=auth.role)
    with pytest.raises(NotFoundError):
        await maps_export_service.build_workbook(db, other, run.id)
