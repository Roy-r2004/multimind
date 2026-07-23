import asyncio
import hashlib
from io import BytesIO
from decimal import Decimal
from pathlib import Path
from datetime import UTC, datetime

import pytest
from httpx import ASGITransport, AsyncClient
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker
from sqlalchemy.orm import selectinload

from app.core.config import get_settings
from app.core.dependencies import AuthContext, get_auth_context
from app.db.models import (
    RehabilitationFacility,
    RehabilitationFacilityAttribute,
    RehabilitationFacilityContact,
    RehabilitationFieldEvidence,
    RehabilitationPossibleDuplicate,
    RehabilitationSource,
    RehabilitationUnresolvedField,
    ScrapingBlueprint,
    ScrapingBlueprintStatus,
    ScrapingCoverageCell,
    ScrapingCoverageStatus,
    ScrapingEvent,
    ScrapingExecution,
    ScrapingExecutionAgent,
    ScrapingExecutionAgentStatus,
    ScrapingExecutionStatus,
    ScrapingMission,
    ScrapingMissionStatus,
    ScrapingRun,
    ScrapingRunAgent,
    ScrapingRunStatus,
    ScrapingSourceCandidate,
    ScrapingSourceDiscoveryQuery,
    ScrapingSourceDocument,
    ScrapingSourceRetrievalAttempt,
    ScrapingTask,
    ScrapingTaskStatus,
    SourceCandidateStatus,
    SourceDiscoveryQueryStatus,
    SourceRetrievalAttemptStatus,
)
from app.llm.providers import LLMResponse
from app.main import create_app
from app.schemas.api import (
    ScrapingBlueprintChangeRequest,
    ScrapingBlueprintContent,
    ScrapingBlueprintRejectRequest,
    ScrapingBlueprintRenameRequest,
    ScrapingExecutionCreate,
    ScrapingMissionCreate,
    ScrapingMissionUpdate,
    ScrapingTeamPlanOutput,
)
from app.services.scraping.source_discovery_service import SourceDiscoverySummary
from app.services.scraping.source_retrieval_service import (
    SourceRetrievalSummary,
)
from app.scraping.worker import WorkerSettings, run_scraping_execution
from app.scraping.blueprint_orchestrator import BlueprintOrchestrator
from app.services.domain_service import project_service
from app.services.scraping.blueprint_service import blueprint_service
from app.services.scraping.countries import COUNTRIES, resolve_country
from app.services.scraping.execution_orchestrator import (
    METRIC_REFRESH_TASK_INTERVAL,
    SourceDiscoveryExecutionOrchestrator,
)
from app.services.scraping.execution_export_service import SHEET_ORDER, safe_cell
from app.services.scraping.execution_outcome import coverage_gap_count, execution_outcome_label
from app.services.scraping.execution_service import execution_service
from app.services.scraping.mock_facility_generator import (
    _CellFallback,
    _facility_coverage_contexts,
    _typed_attribute_values,
    mock_facility_generator,
)
from app.services.scraping.mission_service import mission_service
from app.services.scraping.run_service import run_service
from app.services.scraping.team_planner_service import TeamPlannerService, team_planner_service
from conftest import create_model_set, create_other_auth, create_project, valid_blueprint


class FakeOrchestrator:
    def __init__(self, payload: dict | None = None, should_fail: bool = False) -> None:
        self.payload = payload or valid_blueprint()
        self.should_fail = should_fail
        self.calls = []

    async def generate(self, mission, model_set, previous_blueprint=None, change_instructions=None):
        self.calls.append((mission, model_set, previous_blueprint, change_instructions))
        if self.should_fail:
            raise RuntimeError("provider failed")
        return ScrapingBlueprintContent.model_validate(self.payload)


def logged_execution_reason(caplog, reason: str) -> bool:
    return any(getattr(record, "reason", None) == reason for record in caplog.records)


async def create_mission(db: AsyncSession, auth: AuthContext) -> str:
    return await create_country_mission(db, auth, country_code="LB")


async def create_country_mission(
    db: AsyncSession, auth: AuthContext, *, country_code: str
) -> str:
    await create_model_set(db, auth)
    mission = await mission_service.create_mission(
        db,
        auth,
        ScrapingMissionCreate(
            title="Mission",
            country_code=country_code,
            original_prompt="Find facilities",
            model_set_id="research-set",
        ),
    )
    return mission.id


async def create_blueprint_version(
    db: AsyncSession,
    auth: AuthContext,
    *,
    mission_id: str | None = None,
    version: int = 1,
    status: ScrapingBlueprintStatus = ScrapingBlueprintStatus.DRAFT,
    active: bool = False,
) -> ScrapingBlueprint:
    if mission_id is None:
        mission_id = await create_mission(db, auth)
    blueprint = ScrapingBlueprint(
        mission_id=mission_id,
        version=version,
        status=status,
        blueprint_json=None if status == ScrapingBlueprintStatus.GENERATING else valid_blueprint(),
        model_set_id="research-set",
        judge_model_id="gpt-4.1",
    )
    db.add(blueprint)
    await db.flush()
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    if active:
        mission.active_blueprint_id = blueprint.id
        mission.status = ScrapingMissionStatus.APPROVED
    elif status == ScrapingBlueprintStatus.GENERATING:
        mission.status = ScrapingMissionStatus.BLUEPRINT_GENERATING
    elif status == ScrapingBlueprintStatus.DRAFT:
        mission.status = ScrapingMissionStatus.AWAITING_APPROVAL
    elif status == ScrapingBlueprintStatus.FAILED:
        mission.status = ScrapingMissionStatus.FAILED
    elif status == ScrapingBlueprintStatus.REJECTED:
        mission.status = ScrapingMissionStatus.REJECTED
    await db.flush()
    return blueprint


@pytest.mark.asyncio
async def test_unauthenticated_mission_creation_returns_authentication_error(db: AsyncSession):
    from app.db.session import get_db

    app = create_app()

    async def override_db():
        yield db

    app.dependency_overrides[get_db] = override_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post("/api/v1/scraping/missions", json={})
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_mission_creation_succeeds(db: AsyncSession, auth: AuthContext):
    await create_model_set(db, auth)
    mission = await mission_service.create_mission(
        db,
        auth,
        ScrapingMissionCreate(
            title=" Mission ",
            country_code="LB",
            original_prompt=" Prompt ",
            model_set_id="research-set",
        ),
    )
    assert mission.title == "Mission"
    assert mission.original_prompt == "Prompt"
    assert mission.status == "draft"


@pytest.mark.asyncio
async def test_mission_creation_rejects_empty_title(db: AsyncSession, auth: AuthContext):
    await create_model_set(db, auth)
    with pytest.raises(Exception, match="Mission title is required"):
        await mission_service.create_mission(
            db,
            auth,
            ScrapingMissionCreate(
                title=" ",
                country_code="LB",
                original_prompt="Prompt",
                model_set_id="research-set",
            ),
        )


@pytest.mark.asyncio
async def test_mission_creation_rejects_empty_prompt(db: AsyncSession, auth: AuthContext):
    await create_model_set(db, auth)
    with pytest.raises(Exception, match="Mission prompt is required"):
        await mission_service.create_mission(
            db,
            auth,
            ScrapingMissionCreate(
                title="Mission",
                country_code="LB",
                original_prompt=" ",
                model_set_id="research-set",
            ),
        )


@pytest.mark.asyncio
async def test_mission_creation_rejects_another_organizations_model_set(db: AsyncSession, auth: AuthContext):
    other = await create_other_auth(db)
    await create_model_set(db, other, slug="other-set")
    with pytest.raises(Exception, match="ModelSet not found"):
        await mission_service.create_mission(
            db,
            auth,
            ScrapingMissionCreate(
                title="Mission",
                country_code="LB",
                original_prompt="Prompt",
                model_set_id="other-set",
            ),
        )


@pytest.mark.asyncio
async def test_mission_creation_rejects_another_organizations_project(db: AsyncSession, auth: AuthContext):
    await create_model_set(db, auth)
    other = await create_other_auth(db)
    other_project = await create_project(db, other)
    with pytest.raises(Exception, match="Project not found"):
        await mission_service.create_mission(
            db,
            auth,
            ScrapingMissionCreate(
                title="Mission",
                country_code="LB",
                original_prompt="Prompt",
                model_set_id="research-set",
                project_id=other_project.id,
            ),
        )


@pytest.mark.asyncio
async def test_mission_list_returns_only_current_organization_missions(db: AsyncSession, auth: AuthContext):
    await create_mission(db, auth)
    other = await create_other_auth(db)
    await create_mission(db, other)
    missions = await mission_service.list_missions(db, auth)
    assert len(missions) == 1


@pytest.mark.asyncio
async def test_mission_detail_rejects_another_organization(db: AsyncSession, auth: AuthContext):
    other = await create_other_auth(db)
    mission_id = await create_mission(db, other)
    with pytest.raises(Exception, match="ScrapingMission not found"):
        await mission_service.get_mission(db, auth, mission_id)


@pytest.mark.asyncio
async def test_blueprint_generation_creates_version_1_and_uses_selected_model_set(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    fake = FakeOrchestrator()
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: fake)
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    assert blueprint.version == 1
    assert blueprint.model_set_id == "research-set"
    assert fake.calls[0][1].slug == "research-set"


@pytest.mark.asyncio
async def test_second_blueprint_generation_creates_version_2(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    first = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.reject_blueprint(db, auth, first.id, ScrapingBlueprintRejectRequest(reason="No"))
    second = await blueprint_service.generate_blueprint(db, auth, mission_id)
    assert second.version == 2


@pytest.mark.asyncio
async def test_valid_mocked_judge_json_is_saved(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    assert blueprint.blueprint_json is not None
    assert blueprint.status == "draft"


@pytest.mark.asyncio
async def test_invalid_repair_output_marks_blueprint_failed(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr(
        "app.services.scraping.blueprint_service.get_blueprint_orchestrator",
        lambda: FakeOrchestrator(should_fail=True),
    )
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    assert blueprint.status == "failed"
    assert blueprint.blueprint_json is None


@pytest.mark.asyncio
async def test_invalid_judge_json_triggers_one_repair_call():
    from app.llm.providers import LLMResponse

    class Provider:
        def __init__(self):
            self.calls = 0

        async def complete(self, **kwargs):
            self.calls += 1
            return LLMResponse(text=__import__("json").dumps(valid_blueprint()), tokens_input=1, tokens_output=1)

    provider = Provider()
    orchestrator = BlueprintOrchestrator()
    result = await orchestrator._parse_validate_or_repair(
        provider=provider,
        model="openai/gpt-4.1",
        invalid_output='{"mission_summary": {}}',
    )
    assert result.mission_summary.goal == "Find target entities"
    assert provider.calls == 1


@pytest.mark.asyncio
async def test_approve_draft_blueprint_succeeds_sets_active_status_and_supersedes(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    older = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.reject_blueprint(db, auth, older.id, ScrapingBlueprintRejectRequest(reason="Change"))
    newer = await blueprint_service.generate_blueprint(db, auth, mission_id)
    approved = await blueprint_service.approve_blueprint(db, auth, newer.id)
    mission = await mission_service.get_mission(db, auth, mission_id)
    assert approved.status == "approved"
    assert mission.active_blueprint_id == newer.id
    assert mission.status == "approved"


@pytest.mark.asyncio
async def test_approve_non_draft_blueprint_fails(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.approve_blueprint(db, auth, blueprint.id)
    with pytest.raises(Exception, match="Only draft blueprints can be approved"):
        await blueprint_service.approve_blueprint(db, auth, blueprint.id)


@pytest.mark.asyncio
async def test_approval_supersedes_older_versions_and_does_not_start_scraping(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    first = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.approve_blueprint(db, auth, first.id)
    second = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.approve_blueprint(db, auth, second.id)
    rows = (await db.execute(select(ScrapingBlueprint).order_by(ScrapingBlueprint.version))).scalars().all()
    assert rows[0].status == ScrapingBlueprintStatus.SUPERSEDED
    assert rows[1].status == ScrapingBlueprintStatus.APPROVED
    assert (await db.get(ScrapingMission, mission_id)).status == ScrapingMissionStatus.APPROVED


@pytest.mark.asyncio
async def test_reject_draft_blueprint_succeeds_and_requires_reason(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    with pytest.raises(Exception, match="Rejection reason is required"):
        await blueprint_service.reject_blueprint(db, auth, blueprint.id, ScrapingBlueprintRejectRequest(reason=" "))
    rejected = await blueprint_service.reject_blueprint(db, auth, blueprint.id, ScrapingBlueprintRejectRequest(reason="Bad scope"))
    assert rejected.status == "rejected"


@pytest.mark.asyncio
async def test_request_changes_creates_new_version_preserves_previous_and_does_not_replace_active(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    original = await blueprint_service.generate_blueprint(db, auth, mission_id)
    await blueprint_service.approve_blueprint(db, auth, original.id)
    created = await blueprint_service.request_changes(
        db,
        auth,
        original.id,
        ScrapingBlueprintChangeRequest(change_instructions="Tighten scope"),
    )
    mission = await mission_service.get_mission(db, auth, mission_id)
    assert created.version == 2
    assert created.status == "draft"
    assert mission.active_blueprint_id == original.id
    previous = await blueprint_service.get_blueprint(db, auth, original.id)
    assert previous.status == "approved"


@pytest.mark.asyncio
async def test_another_organization_cannot_approve_a_blueprint(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingBlueprint not found"):
        await blueprint_service.approve_blueprint(db, other, blueprint.id)


@pytest.mark.asyncio
async def test_approved_blueprint_cannot_be_modified(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    blueprint = await blueprint_service.generate_blueprint(db, auth, mission_id)
    approved = await blueprint_service.approve_blueprint(db, auth, blueprint.id)
    with pytest.raises(Exception, match="Only draft blueprints can be rejected"):
        await blueprint_service.reject_blueprint(db, auth, approved.id, ScrapingBlueprintRejectRequest(reason="No"))


@pytest.mark.asyncio
async def test_mission_deletion_deletes_blueprint_history(db: AsyncSession, auth: AuthContext, monkeypatch):
    mission_id = await create_mission(db, auth)
    monkeypatch.setattr("app.services.scraping.blueprint_service.get_blueprint_orchestrator", lambda: FakeOrchestrator())
    await blueprint_service.generate_blueprint(db, auth, mission_id)
    await mission_service.delete_mission(db, auth, mission_id)
    rows = (await db.execute(select(ScrapingBlueprint))).scalars().all()
    assert rows == []


@pytest.mark.asyncio
async def test_rename_draft_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth)
    renamed = await blueprint_service.rename_blueprint(
        db, auth, blueprint.id, ScrapingBlueprintRenameRequest(name="Final Lebanon Strategy")
    )
    assert renamed.display_name == "Final Lebanon Strategy"
    assert renamed.version == 1
    assert renamed.status == "draft"


@pytest.mark.asyncio
async def test_rename_approved_blueprint_succeeds_without_changing_blueprint_json(
    db: AsyncSession, auth: AuthContext
):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.APPROVED, active=True
    )
    before = dict(blueprint.blueprint_json)
    renamed = await blueprint_service.rename_blueprint(
        db, auth, blueprint.id, ScrapingBlueprintRenameRequest(name="Approved Strategy")
    )
    row = await db.get(ScrapingBlueprint, blueprint.id)
    assert renamed.display_name == "Approved Strategy"
    assert row is not None
    assert row.blueprint_json == before
    assert row.status == ScrapingBlueprintStatus.APPROVED


@pytest.mark.asyncio
async def test_rename_rejected_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.REJECTED
    )
    renamed = await blueprint_service.rename_blueprint(
        db, auth, blueprint.id, ScrapingBlueprintRenameRequest(name="Rejected Strategy")
    )
    assert renamed.display_name == "Rejected Strategy"
    assert renamed.status == "rejected"


@pytest.mark.asyncio
async def test_rename_trims_whitespace(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth)
    renamed = await blueprint_service.rename_blueprint(
        db, auth, blueprint.id, ScrapingBlueprintRenameRequest(name="  Trimmed Name  ")
    )
    assert renamed.display_name == "Trimmed Name"


def test_rename_rejects_empty_name():
    with pytest.raises(Exception):
        ScrapingBlueprintRenameRequest(name="")


def test_rename_rejects_whitespace_only_name():
    with pytest.raises(Exception):
        ScrapingBlueprintRenameRequest(name="   ")


def test_rename_rejects_names_longer_than_160_characters():
    with pytest.raises(Exception):
        ScrapingBlueprintRenameRequest(name="x" * 161)


@pytest.mark.asyncio
async def test_rename_generating_blueprint_fails(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.GENERATING
    )
    with pytest.raises(Exception, match="currently generating cannot be renamed"):
        await blueprint_service.rename_blueprint(
            db, auth, blueprint.id, ScrapingBlueprintRenameRequest(name="Generating")
        )


@pytest.mark.asyncio
async def test_another_organization_cannot_rename_a_blueprint(
    db: AsyncSession, auth: AuthContext
):
    blueprint = await create_blueprint_version(db, auth)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingBlueprint not found"):
        await blueprint_service.rename_blueprint(
            db, other, blueprint.id, ScrapingBlueprintRenameRequest(name="No Access")
        )


@pytest.mark.asyncio
async def test_delete_draft_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth)
    await blueprint_service.delete_blueprint(db, auth, blueprint.id)
    assert await db.get(ScrapingBlueprint, blueprint.id) is None


@pytest.mark.asyncio
async def test_delete_rejected_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.REJECTED
    )
    await blueprint_service.delete_blueprint(db, auth, blueprint.id)
    assert await db.get(ScrapingBlueprint, blueprint.id) is None


@pytest.mark.asyncio
async def test_delete_superseded_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.SUPERSEDED
    )
    await blueprint_service.delete_blueprint(db, auth, blueprint.id)
    assert await db.get(ScrapingBlueprint, blueprint.id) is None


@pytest.mark.asyncio
async def test_delete_failed_blueprint_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth, status=ScrapingBlueprintStatus.FAILED)
    await blueprint_service.delete_blueprint(db, auth, blueprint.id)
    assert await db.get(ScrapingBlueprint, blueprint.id) is None


@pytest.mark.asyncio
async def test_delete_approved_blueprint_fails(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.APPROVED
    )
    with pytest.raises(Exception, match="active approved blueprint cannot be deleted"):
        await blueprint_service.delete_blueprint(db, auth, blueprint.id)


@pytest.mark.asyncio
async def test_delete_active_blueprint_fails(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth, active=True)
    with pytest.raises(Exception, match="active approved blueprint cannot be deleted"):
        await blueprint_service.delete_blueprint(db, auth, blueprint.id)


@pytest.mark.asyncio
async def test_delete_generating_blueprint_fails(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.GENERATING
    )
    with pytest.raises(Exception, match="currently generating cannot be deleted"):
        await blueprint_service.delete_blueprint(db, auth, blueprint.id)


@pytest.mark.asyncio
async def test_another_organization_cannot_delete_a_blueprint(
    db: AsyncSession, auth: AuthContext
):
    blueprint = await create_blueprint_version(db, auth)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingBlueprint not found"):
        await blueprint_service.delete_blueprint(db, other, blueprint.id)


@pytest.mark.asyncio
async def test_deleting_one_version_preserves_other_versions(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    first = await create_blueprint_version(db, auth, mission_id=mission_id, version=1)
    second = await create_blueprint_version(db, auth, mission_id=mission_id, version=2)
    third = await create_blueprint_version(db, auth, mission_id=mission_id, version=3)
    await blueprint_service.delete_blueprint(db, auth, second.id)
    rows = (
        (await db.execute(select(ScrapingBlueprint).order_by(ScrapingBlueprint.version)))
        .scalars()
        .all()
    )
    assert [row.id for row in rows] == [first.id, third.id]


@pytest.mark.asyncio
async def test_deleting_a_version_does_not_renumber_remaining_versions(
    db: AsyncSession, auth: AuthContext
):
    mission_id = await create_mission(db, auth)
    await create_blueprint_version(db, auth, mission_id=mission_id, version=1)
    second = await create_blueprint_version(db, auth, mission_id=mission_id, version=2)
    await create_blueprint_version(db, auth, mission_id=mission_id, version=3)
    await blueprint_service.delete_blueprint(db, auth, second.id)
    rows = (
        (await db.execute(select(ScrapingBlueprint).order_by(ScrapingBlueprint.version)))
        .scalars()
        .all()
    )
    assert [row.version for row in rows] == [1, 3]


@pytest.mark.asyncio
async def test_deleting_the_final_blueprint_sets_mission_status_draft(
    db: AsyncSession, auth: AuthContext
):
    blueprint = await create_blueprint_version(db, auth)
    mission_id = blueprint.mission_id
    await blueprint_service.delete_blueprint(db, auth, blueprint.id)
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    assert mission.status == ScrapingMissionStatus.DRAFT


@pytest.mark.asyncio
async def test_deleting_a_draft_while_an_active_blueprint_exists_keeps_mission_status_approved(
    db: AsyncSession, auth: AuthContext
):
    mission_id = await create_mission(db, auth)
    await create_blueprint_version(
        db,
        auth,
        mission_id=mission_id,
        version=1,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    draft = await create_blueprint_version(db, auth, mission_id=mission_id, version=2)
    await blueprint_service.delete_blueprint(db, auth, draft.id)
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    assert mission.status == ScrapingMissionStatus.APPROVED


@pytest.mark.asyncio
async def test_deleting_the_selected_draft_does_not_clear_an_unrelated_active_blueprint_id(
    db: AsyncSession, auth: AuthContext
):
    mission_id = await create_mission(db, auth)
    active = await create_blueprint_version(
        db,
        auth,
        mission_id=mission_id,
        version=1,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    draft = await create_blueprint_version(db, auth, mission_id=mission_id, version=2)
    await blueprint_service.delete_blueprint(db, auth, draft.id)
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    assert mission.active_blueprint_id == active.id


@pytest.mark.asyncio
async def test_explicit_project_id_null_removes_the_mission_from_its_project(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    updated = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=None)
    )
    assert updated.project_id is None


@pytest.mark.asyncio
async def test_assigning_a_mission_to_a_project_succeeds(db: AsyncSession, auth: AuthContext):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    updated = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    assert updated.project_id == project.id
    assert updated.project_name == project.name


@pytest.mark.asyncio
async def test_project_assignment_persists_after_reloading_the_mission(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    reloaded = await mission_service.get_mission(db, auth, mission_id)
    assert reloaded.project_id == project.id


@pytest.mark.asyncio
async def test_moving_from_one_project_to_another_succeeds(
    db: AsyncSession, auth: AuthContext
):
    old_project = await create_project(db, auth, name="Old Project")
    new_project = await create_project(db, auth, name="New Project")
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=old_project.id)
    )
    moved = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=new_project.id)
    )
    assert moved.project_id == new_project.id


@pytest.mark.asyncio
async def test_assigning_a_project_does_not_change_blueprint_statuses(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    blueprint = await create_blueprint_version(db, auth)
    await mission_service.update_mission(
        db, auth, blueprint.mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    row = await db.get(ScrapingBlueprint, blueprint.id)
    assert row is not None
    assert row.status == ScrapingBlueprintStatus.DRAFT


@pytest.mark.asyncio
async def test_assigning_a_project_does_not_create_a_new_blueprint(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    blueprint = await create_blueprint_version(db, auth)
    await mission_service.update_mission(
        db, auth, blueprint.mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    rows = (
        (
            await db.execute(
                select(ScrapingBlueprint).where(
                    ScrapingBlueprint.mission_id == blueprint.mission_id
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(rows) == 1


@pytest.mark.asyncio
async def test_omitted_project_id_preserves_the_current_project(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    updated = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(title="Renamed Mission")
    )
    assert updated.project_id == project.id


@pytest.mark.asyncio
async def test_assigning_another_organizations_project_fails(
    db: AsyncSession, auth: AuthContext
):
    other = await create_other_auth(db)
    other_project = await create_project(db, other)
    mission_id = await create_mission(db, auth)
    with pytest.raises(Exception, match="Project not found"):
        await mission_service.update_mission(
            db, auth, mission_id, ScrapingMissionUpdate(project_id=other_project.id)
        )


@pytest.mark.asyncio
async def test_project_detail_returns_assigned_scraping_missions(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    detail = await project_service.get_detail(db, auth, project.id)
    assert [mission.id for mission in detail.scraping_missions] == [mission_id]
    assert not hasattr(detail.scraping_missions[0], "blueprint_json")


@pytest.mark.asyncio
async def test_project_detail_does_not_return_missions_from_other_projects(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth, name="Project A")
    other_project = await create_project(db, auth, name="Project B")
    mission_id = await create_mission(db, auth)
    other_mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    await mission_service.update_mission(
        db, auth, other_mission_id, ScrapingMissionUpdate(project_id=other_project.id)
    )
    detail = await project_service.get_detail(db, auth, project.id)
    assert [mission.id for mission in detail.scraping_missions] == [mission_id]


@pytest.mark.asyncio
async def test_moving_a_mission_removes_it_from_the_old_project_response(
    db: AsyncSession, auth: AuthContext
):
    old_project = await create_project(db, auth, name="Old Project")
    new_project = await create_project(db, auth, name="New Project")
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=old_project.id)
    )
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=new_project.id)
    )
    old_detail = await project_service.get_detail(db, auth, old_project.id)
    new_detail = await project_service.get_detail(db, auth, new_project.id)
    assert old_detail.scraping_missions == []
    assert [mission.id for mission in new_detail.scraping_missions] == [mission_id]


@pytest.mark.asyncio
async def test_removing_project_id_removes_mission_from_project_detail(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=None)
    )
    detail = await project_service.get_detail(db, auth, project.id)
    assert detail.scraping_missions == []


@pytest.mark.asyncio
async def test_cross_organization_missions_are_not_exposed_in_project_detail(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    other = await create_other_auth(db)
    other_project = await create_project(db, other)
    other_mission_id = await create_mission(db, other)
    await mission_service.update_mission(
        db, other, other_mission_id, ScrapingMissionUpdate(project_id=other_project.id)
    )
    detail = await project_service.get_detail(db, auth, project.id)
    assert detail.scraping_missions == []


@pytest.mark.asyncio
async def test_mission_rename_succeeds(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    updated = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(title="New Mission Name")
    )
    assert updated.title == "New Mission Name"


@pytest.mark.asyncio
async def test_mission_rename_trims_whitespace(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    updated = await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(title="  Trimmed Mission  ")
    )
    assert updated.title == "Trimmed Mission"


@pytest.mark.asyncio
async def test_mission_rename_rejects_empty_title(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    with pytest.raises(Exception, match="Mission title is required"):
        await mission_service.update_mission(
            db, auth, mission_id, ScrapingMissionUpdate(title="   ")
        )


@pytest.mark.asyncio
async def test_mission_rename_is_reflected_in_project_detail(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(title="Project Visible Name")
    )
    detail = await project_service.get_detail(db, auth, project.id)
    assert detail.scraping_missions[0].title == "Project Visible Name"


@pytest.mark.asyncio
async def test_cross_organization_mission_rename_fails(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingMission not found"):
        await mission_service.update_mission(
            db, other, mission_id, ScrapingMissionUpdate(title="No Access")
        )


@pytest.mark.asyncio
async def test_delete_draft_mission_succeeds(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    await mission_service.delete_mission(db, auth, mission_id)
    assert await db.get(ScrapingMission, mission_id) is None


@pytest.mark.asyncio
async def test_delete_approved_mission_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.APPROVED, active=True
    )
    await mission_service.delete_mission(db, auth, blueprint.mission_id)
    assert await db.get(ScrapingMission, blueprint.mission_id) is None


@pytest.mark.asyncio
async def test_delete_rejected_mission_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(
        db, auth, status=ScrapingBlueprintStatus.REJECTED
    )
    await mission_service.delete_mission(db, auth, blueprint.mission_id)
    assert await db.get(ScrapingMission, blueprint.mission_id) is None


@pytest.mark.asyncio
async def test_delete_failed_mission_succeeds(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth, status=ScrapingBlueprintStatus.FAILED)
    await mission_service.delete_mission(db, auth, blueprint.mission_id)
    assert await db.get(ScrapingMission, blueprint.mission_id) is None


@pytest.mark.asyncio
async def test_deleting_a_mission_removes_its_blueprints(db: AsyncSession, auth: AuthContext):
    blueprint = await create_blueprint_version(db, auth)
    await mission_service.delete_mission(db, auth, blueprint.mission_id)
    rows = (await db.execute(select(ScrapingBlueprint))).scalars().all()
    assert rows == []


@pytest.mark.asyncio
async def test_deleting_a_mission_removes_it_from_project_detail(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    await mission_service.delete_mission(db, auth, mission_id)
    detail = await project_service.get_detail(db, auth, project.id)
    assert detail.scraping_missions == []


@pytest.mark.asyncio
async def test_deleting_a_mission_does_not_delete_the_project(
    db: AsyncSession, auth: AuthContext
):
    project = await create_project(db, auth)
    mission_id = await create_mission(db, auth)
    await mission_service.update_mission(
        db, auth, mission_id, ScrapingMissionUpdate(project_id=project.id)
    )
    await mission_service.delete_mission(db, auth, mission_id)
    assert await db.get(type(project), project.id) is not None


@pytest.mark.asyncio
async def test_deleting_a_mission_does_not_remove_other_missions(
    db: AsyncSession, auth: AuthContext
):
    first = await create_mission(db, auth)
    second = await create_mission(db, auth)
    await mission_service.delete_mission(db, auth, first)
    assert await db.get(ScrapingMission, second) is not None


@pytest.mark.asyncio
async def test_deleting_a_generating_mission_fails(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    mission.status = ScrapingMissionStatus.BLUEPRINT_GENERATING
    await db.flush()
    with pytest.raises(Exception, match="cannot be deleted while its blueprint is generating"):
        await mission_service.delete_mission(db, auth, mission_id)


@pytest.mark.asyncio
async def test_cross_organization_mission_deletion_fails(db: AsyncSession, auth: AuthContext):
    mission_id = await create_mission(db, auth)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingMission not found"):
        await mission_service.delete_mission(db, other, mission_id)


def team_plan_payload(count: int, *, model_id: str = "gpt-4.1") -> dict:
    agents = []
    for index in range(1, count + 1):
        agents.append(
            {
                "sequence": index,
                "name": f"Agent {index}",
                "role": "verification" if index == count else f"role_{index}",
                "purpose": f"Purpose {index}",
                "instructions": f"Instructions {index}",
                "assigned_scope": {
                    "regions": [f"Region {index}"],
                    "languages": ["en"],
                    "source_categories": ["official"],
                },
                "model_id": model_id,
                "depends_on": [index - 1] if index > 1 else [],
            }
        )
    return {
        "recommended_agent_count": count,
        "rationale": f"{count} agents are appropriate for the approved blueprint.",
        "agents": agents,
    }


async def create_active_approved_blueprint(
    db: AsyncSession, auth: AuthContext
) -> ScrapingBlueprint:
    return await create_blueprint_version(
        db,
        auth,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )


def mock_planner(monkeypatch, count: int = 3):
    async def fake_plan_team(mission, blueprint, model_set):
        return (
            ScrapingTeamPlanOutput.model_validate(team_plan_payload(count)),
            model_set.verdict_model,
        )

    monkeypatch.setattr(team_planner_service, "plan_team", fake_plan_team)


def mock_source_discovery(monkeypatch, candidate_count: int = 1) -> None:
    async def fake_discover(service_db, context):
        now = datetime.now(UTC)
        query = ScrapingSourceDiscoveryQuery(
            organization_id=context.organization_id,
            execution_id=context.execution_id,
            coverage_cell_id=context.coverage_cell_id,
            task_id=context.task_id,
            country_code=context.country_code,
            country_name=context.country_name,
            region_code=context.region_code,
            region_name=context.region_name,
            language_code=context.language_code,
            language_name=context.language_name,
            source_category=context.source_category,
            query_text="real source discovery query",
            provider="serper",
            status=SourceDiscoveryQueryStatus.SUCCEEDED,
            requested_at=now,
            completed_at=now,
            result_count=candidate_count,
            metadata_json={"purpose": "test"},
        )
        service_db.add(query)
        await service_db.flush()
        for index in range(candidate_count):
            url = f"https://sources.example.org/{context.task_id}/{index + 1}"
            service_db.add(
                ScrapingSourceCandidate(
                    organization_id=context.organization_id,
                    execution_id=context.execution_id,
                    coverage_cell_id=context.coverage_cell_id,
                    discovery_query_id=query.id,
                    provider="serper",
                    provider_result_id=url,
                    rank=index + 1,
                    url=url,
                    canonical_url=url,
                    domain="sources.example.org",
                    title=f"Real source {index + 1}",
                    snippet="Real snippet",
                    country_code=context.country_code,
                    country_name=context.country_name,
                    region_code=context.region_code,
                    region_name=context.region_name,
                    language_code=context.language_code,
                    language_name=context.language_name,
                    source_category=context.source_category,
                    initial_relevance_score=Decimal("1.0"),
                    initial_trust_tier="high",
                    status=SourceCandidateStatus.DISCOVERED,
                    discovered_at=now,
                    metadata_json={"position": index + 1},
                )
            )
        await service_db.flush()
        return SourceDiscoverySummary(
            provider="serper",
            planned_query_count=1,
            query_count=1,
            succeeded_query_count=1,
            failed_query_count=0,
            candidate_count=candidate_count,
            duplicate_candidate_count=0,
            rejected_result_count=0,
        )

    monkeypatch.setattr(
        "app.services.scraping.execution_orchestrator.source_discovery_service.discover",
        fake_discover,
    )

    async def fake_retrieve(service_db, context):
        content = b"real public health source"
        content_hash = hashlib.sha256(content).hexdigest()
        attempt = ScrapingSourceRetrievalAttempt(
            organization_id=context.organization_id,
            execution_id=context.execution_id,
            source_candidate_id=context.source_candidate_id,
            coverage_cell_id=context.coverage_cell_id,
            task_id=context.task_id,
            status=SourceRetrievalAttemptStatus.SUCCEEDED,
            requested_url="https://sante.gouv.fr/source",
            final_url="https://sante.gouv.fr/source",
            redirect_count=0,
            http_status=200,
            content_type="text/html",
            bytes_received=len(content),
            started_at=datetime.now(UTC),
            completed_at=datetime.now(UTC),
            idempotency_key=context.idempotency_key,
            metadata_json={"content_sha256": content_hash},
        )
        service_db.add(attempt)
        await service_db.flush()
        document = ScrapingSourceDocument(
            organization_id=context.organization_id,
            execution_id=context.execution_id,
            source_candidate_id=context.source_candidate_id,
            retrieval_attempt_id=attempt.id,
            final_url="https://sante.gouv.fr/source",
            content_type="text/html",
            content_sha256=content_hash,
            content_text=content.decode(),
            byte_size=len(content),
            retrieval_timestamp=datetime.now(UTC),
            metadata_json={"test": "orchestrator"},
        )
        service_db.add(document)
        await service_db.flush()
        return SourceRetrievalSummary(
            attempt_id=attempt.id,
            status="succeeded",
            requested_url=attempt.requested_url,
            final_url=attempt.final_url,
            redirect_count=0,
            http_status=200,
            content_type="text/html",
            bytes_received=len(content),
            robots_status=None,
            document_id=document.id,
            content_sha256=content_hash,
        )

    monkeypatch.setattr(
        "app.services.scraping.execution_orchestrator.source_retrieval_service.retrieve",
        fake_retrieve,
    )


@pytest.mark.asyncio
@pytest.mark.parametrize("count", [3, 7])
async def test_dynamic_planning_persists_ai_selected_agent_count(
    db: AsyncSession, auth: AuthContext, monkeypatch, count: int
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, count)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    assert run.status == "planned"
    assert run.recommended_agent_count == count
    assert len(run.agents) == count
    assert count != 5 or len(run.agents) == 5
    assert run.blueprint_id == blueprint.id
    assert run.blueprint_version == blueprint.version
    assert run.planner_model_id == "gpt-4.1"
    assert run.planner_rationale == f"{count} agents are appropriate for the approved blueprint."
    assert [agent.sequence for agent in run.agents] == list(range(1, count + 1))
    assert run.agents[1].dependency_agent_ids == [run.agents[0].id]
    assert run.agents[0].assigned_scope["regions"] == ["Region 1"]


@pytest.mark.asyncio
async def test_planning_rejects_mission_without_active_approved_blueprint(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_blueprint_version(db, auth, status=ScrapingBlueprintStatus.DRAFT)
    mock_planner(monkeypatch)
    with pytest.raises(Exception, match="active approved blueprint"):
        await run_service.plan_team(db, auth, blueprint.mission_id)


@pytest.mark.asyncio
async def test_planning_rejects_generating_mission(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_blueprint_version(
        db,
        auth,
        status=ScrapingBlueprintStatus.GENERATING,
    )
    mock_planner(monkeypatch)
    with pytest.raises(Exception, match="active approved blueprint"):
        await run_service.plan_team(db, auth, blueprint.mission_id)


@pytest.mark.asyncio
async def test_planning_does_not_use_non_active_approved_blueprint(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    mission_id = await create_mission(db, auth)
    approved = await create_blueprint_version(
        db,
        auth,
        mission_id=mission_id,
        status=ScrapingBlueprintStatus.APPROVED,
    )
    draft = await create_blueprint_version(
        db,
        auth,
        mission_id=mission_id,
        version=2,
        status=ScrapingBlueprintStatus.DRAFT,
        active=True,
    )
    mission = await db.get(ScrapingMission, mission_id)
    assert mission is not None
    mission.active_blueprint_id = draft.id
    await db.flush()
    mock_planner(monkeypatch)
    with pytest.raises(Exception, match="active approved blueprint"):
        await run_service.plan_team(db, auth, approved.mission_id)


@pytest.mark.asyncio
async def test_cross_organization_mission_cannot_plan_run(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    other = await create_other_auth(db)
    mock_planner(monkeypatch)
    with pytest.raises(Exception, match="ScrapingMission not found"):
        await run_service.plan_team(db, other, blueprint.mission_id)


@pytest.mark.asyncio
async def test_duplicate_planning_run_is_rejected(db: AsyncSession, auth: AuthContext, monkeypatch):
    blueprint = await create_active_approved_blueprint(db, auth)
    db.add(
        ScrapingRun(
            organization_id=auth.org_id,
            mission_id=blueprint.mission_id,
            blueprint_id=blueprint.id,
            model_set_id="research-set",
            status=ScrapingRunStatus.PLANNING,
        )
    )
    await db.flush()
    mock_planner(monkeypatch)
    with pytest.raises(Exception, match="already exists"):
        await run_service.plan_team(db, auth, blueprint.mission_id)


@pytest.mark.asyncio
async def test_second_plan_for_exact_blueprint_returns_409(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    from app.db.session import get_db

    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch)
    app = create_app()

    async def override_db():
        yield db

    async def override_auth():
        return auth

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_auth_context] = override_auth
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        first = await client.post(f"/api/v1/scraping/missions/{blueprint.mission_id}/runs/plan")
        second = await client.post(f"/api/v1/scraping/missions/{blueprint.mission_id}/runs/plan")

    assert first.status_code == 200
    assert second.status_code == 409
    body = second.json()
    assert body["message"] == "An AI scraping team plan already exists for this blueprint version."
    assert body["details"] == {
        "message": "An AI scraping team plan already exists for this blueprint version.",
        "existing_run_id": first.json()["id"],
        "existing_run_status": "planned",
    }


@pytest.mark.asyncio
async def test_different_approved_blueprint_version_can_receive_own_run(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    first_blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch)
    first_run = await run_service.plan_team(db, auth, first_blueprint.mission_id)
    second_blueprint = await create_blueprint_version(
        db,
        auth,
        mission_id=first_blueprint.mission_id,
        version=2,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    second_run = await run_service.plan_team(db, auth, first_blueprint.mission_id)
    assert first_run.blueprint_id == first_blueprint.id
    assert second_run.blueprint_id == second_blueprint.id
    assert first_run.id != second_run.id


@pytest.mark.parametrize(
    ("payload", "message"),
    [
        (team_plan_payload(0), "too few|number of agents|at least 1"),
        (team_plan_payload(1), "too few"),
        (team_plan_payload(13), "too many"),
        ({**team_plan_payload(3), "recommended_agent_count": 2}, "recommended_agent_count"),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {**team_plan_payload(3)["agents"][0], "sequence": 1},
                    {**team_plan_payload(3)["agents"][1], "sequence": 1},
                    team_plan_payload(3)["agents"][2],
                ],
            },
            "duplicate",
        ),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {**agent, "model_id": "invented"}
                    for agent in team_plan_payload(3)["agents"]
                ],
            },
            "outside",
        ),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {
                        key: value
                        for key, value in team_plan_payload(3)["agents"][0].items()
                        if key != "role"
                    },
                    team_plan_payload(3)["agents"][1],
                    team_plan_payload(3)["agents"][2],
                ],
            },
            "schema",
        ),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {**team_plan_payload(3)["agents"][0], "depends_on": [1]},
                    team_plan_payload(3)["agents"][1],
                    team_plan_payload(3)["agents"][2],
                ],
            },
            "self-dependency",
        ),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {**team_plan_payload(3)["agents"][0], "depends_on": [99]},
                    team_plan_payload(3)["agents"][1],
                    team_plan_payload(3)["agents"][2],
                ],
            },
            "unknown dependency",
        ),
        (
            {
                **team_plan_payload(3),
                "agents": [
                    {**team_plan_payload(3)["agents"][0], "depends_on": [3]},
                    team_plan_payload(3)["agents"][1],
                    {**team_plan_payload(3)["agents"][2], "depends_on": [1]},
                ],
            },
            "cycle",
        ),
        ({**team_plan_payload(3), "unexpected": "field"}, "schema"),
    ],
)
def test_team_plan_validation_rejects_invalid_structures(payload, message):
    with pytest.raises(Exception, match=message):
        TeamPlannerService().validate_plan_data(payload, allowed_model_ids=["gpt-4.1", "claude"])


@pytest.mark.asyncio
async def test_malformed_planner_json_triggers_one_repair_call():
    class Provider:
        def __init__(self):
            self.calls = 0

        async def complete(self, **kwargs):
            self.calls += 1
            return LLMResponse(
                text=__import__("json").dumps(team_plan_payload(3)),
                tokens_input=1,
                tokens_output=1,
            )

    provider = Provider()
    plan = await TeamPlannerService().parse_validate_or_repair(
        provider=provider,
        provider_model="openai/gpt-4.1",
        raw_text="{not-json",
        allowed_model_ids=["gpt-4.1", "claude"],
    )
    assert provider.calls == 1
    assert plan.recommended_agent_count == 3


@pytest.mark.asyncio
async def test_invalid_repaired_planner_output_leaves_run_failed(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)

    async def fake_plan_team(mission, blueprint, model_set):
        raise RuntimeError("OpenRouter secret sk-test should not leak")

    monkeypatch.setattr(team_planner_service, "plan_team", fake_plan_team)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    assert run.status == "failed"
    assert run.agents == []
    assert "sk-test" not in (run.error_message or "")


@pytest.mark.asyncio
async def test_list_runs_is_mission_and_organization_scoped(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    first = await create_active_approved_blueprint(db, auth)
    second_mission = ScrapingMission(
        org_id=auth.org_id,
        created_by=auth.user.id,
        model_set_id="research-set",
        title="Second mission",
        original_prompt="Second prompt",
        status=ScrapingMissionStatus.APPROVED,
    )
    db.add(second_mission)
    await db.flush()
    second = ScrapingBlueprint(
        mission_id=second_mission.id,
        version=1,
        status=ScrapingBlueprintStatus.APPROVED,
        blueprint_json=valid_blueprint(),
        model_set_id="research-set",
        judge_model_id="gpt-4.1",
    )
    db.add(second)
    await db.flush()
    second_mission.active_blueprint_id = second.id
    await db.flush()
    mock_planner(monkeypatch, 3)
    first_run = await run_service.plan_team(db, auth, first.mission_id)
    await run_service.plan_team(db, auth, second.mission_id)
    runs = await run_service.list_runs(db, auth, first.mission_id)
    assert [run.id for run in runs] == [first_run.id]
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingMission not found"):
        await run_service.list_runs(db, other, first.mission_id)


@pytest.mark.asyncio
async def test_run_detail_returns_agents_and_rejects_other_organization(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    detail = await run_service.get_run(db, auth, run.id)
    assert len(detail.agents) == 3
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingRun not found"):
        await run_service.get_run(db, other, run.id)


@pytest.mark.asyncio
async def test_cancel_planned_run_preserves_agents_and_history(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    cancelled = await run_service.cancel_run(db, auth, run.id)
    assert cancelled.status == "cancelled"
    assert len(cancelled.agents) == 3
    assert {agent.status for agent in cancelled.agents} == {"cancelled"}
    visible = await run_service.list_runs(db, auth, blueprint.mission_id)
    assert visible[0].id == run.id


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "status",
    [
        ScrapingRunStatus.PLANNED,
        ScrapingRunStatus.COMPLETED,
        ScrapingRunStatus.FAILED,
        ScrapingRunStatus.CANCELLED,
    ],
)
async def test_delete_safe_run_statuses_succeeds(
    db: AsyncSession,
    auth: AuthContext,
    monkeypatch,
    status: ScrapingRunStatus,
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    row = await run_service.get_run_row(db, auth, run.id)
    row.status = status
    await db.flush()
    await run_service.delete_run(db, auth, run.id)
    assert await db.get(ScrapingRun, run.id) is None


@pytest.mark.asyncio
@pytest.mark.parametrize("status", [ScrapingRunStatus.PLANNING, ScrapingRunStatus.RUNNING])
async def test_delete_active_run_statuses_returns_409(
    db: AsyncSession,
    auth: AuthContext,
    status: ScrapingRunStatus,
):
    blueprint = await create_active_approved_blueprint(db, auth)
    run = ScrapingRun(
        organization_id=auth.org_id,
        mission_id=blueprint.mission_id,
        blueprint_id=blueprint.id,
        model_set_id="research-set",
        status=status,
    )
    db.add(run)
    await db.flush()
    with pytest.raises(Exception, match="cannot be deleted while planning or executing"):
        await run_service.delete_run(db, auth, run.id)
    assert await db.get(ScrapingRun, run.id) is not None


@pytest.mark.asyncio
async def test_after_deleting_run_same_blueprint_can_be_planned_again(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    first = await run_service.plan_team(db, auth, blueprint.mission_id)
    await run_service.delete_run(db, auth, first.id)
    second = await run_service.plan_team(db, auth, blueprint.mission_id)
    assert second.blueprint_id == blueprint.id
    assert second.id != first.id


@pytest.mark.asyncio
async def test_deleting_run_removes_agents(db: AsyncSession, auth: AuthContext, monkeypatch):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    await run_service.delete_run(db, auth, run.id)
    rows = await db.execute(select(ScrapingRunAgent).where(ScrapingRunAgent.run_id == run.id))
    assert rows.scalars().all() == []


@pytest.mark.asyncio
async def test_cross_organization_run_deletion_fails(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingRun not found"):
        await run_service.delete_run(db, other, run.id)
    assert await db.get(ScrapingRun, run.id) is not None


@pytest.mark.asyncio
async def test_failed_and_completed_cancellation_rules(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)

    async def fake_plan_team(mission, blueprint, model_set):
        raise RuntimeError("provider failed")

    monkeypatch.setattr(team_planner_service, "plan_team", fake_plan_team)
    failed = await run_service.plan_team(db, auth, blueprint.mission_id)
    same = await run_service.cancel_run(db, auth, failed.id)
    assert same.status == "failed"
    row = await run_service.get_run_row(db, auth, failed.id)
    row.status = ScrapingRunStatus.COMPLETED
    await db.flush()
    with pytest.raises(Exception, match="completed"):
        await run_service.cancel_run(db, auth, failed.id)


@pytest.mark.asyncio
async def test_historical_integrity_and_no_blueprint_or_project_mutation(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    project = await create_project(db, auth)
    blueprint = await create_active_approved_blueprint(db, auth)
    await mission_service.update_mission(
        db,
        auth,
        blueprint.mission_id,
        ScrapingMissionUpdate(project_id=project.id),
    )
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    await create_blueprint_version(
        db,
        auth,
        mission_id=blueprint.mission_id,
        version=2,
        status=ScrapingBlueprintStatus.DRAFT,
    )
    reloaded_blueprint = await db.get(ScrapingBlueprint, blueprint.id)
    assert reloaded_blueprint is not None
    assert run.blueprint_id == blueprint.id
    assert reloaded_blueprint.status == ScrapingBlueprintStatus.APPROVED
    mission = await db.get(ScrapingMission, blueprint.mission_id)
    assert mission is not None
    assert mission.project_id == project.id
    count = await db.execute(
        select(ScrapingBlueprint).where(ScrapingBlueprint.mission_id == blueprint.mission_id)
    )
    assert len(count.scalars().all()) == 2


@pytest.mark.asyncio
async def test_planning_persists_allowed_models_only(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    assert {agent.model_id for agent in run.agents} == {"gpt-4.1"}
    rows = await db.execute(select(ScrapingRunAgent).where(ScrapingRunAgent.run_id == run.id))
    assert len(rows.scalars().all()) == 3


def test_offline_country_catalog_has_complete_valid_iso_codes():
    assert len(COUNTRIES) == 249
    assert len(set(COUNTRIES)) == 249
    assert all(len(code) == 2 for code in COUNTRIES)
    assert all(code.isalpha() and code.isupper() for code in COUNTRIES)
    assert all(country.code == code for code, country in COUNTRIES.items())
    assert all(country.name.strip() for country in COUNTRIES.values())


def test_country_resolution_normalizes_and_rejects_unsupported_codes():
    lebanon = resolve_country("lb")
    assert lebanon.code == "LB"
    assert lebanon.name == "Lebanon"

    turkiye = resolve_country(" tr ")
    assert turkiye.code == "TR"
    assert turkiye.name == "Türkiye"

    aland = resolve_country("ax")
    assert aland.code == "AX"
    assert aland.name == "Åland Islands"

    with pytest.raises(Exception, match="Unsupported country code"):
        resolve_country("XX")
    with pytest.raises(Exception, match="Unsupported country code"):
        resolve_country("XK")


@pytest.mark.asyncio
async def test_mission_creation_requires_valid_country(db: AsyncSession, auth: AuthContext):
    await create_model_set(db, auth)
    with pytest.raises(Exception, match="Unsupported country code"):
        await mission_service.create_mission(
            db,
            auth,
            ScrapingMissionCreate(
                title="Mission",
                country_code="XX",
                original_prompt="Prompt",
                model_set_id="research-set",
            ),
        )
    mission = await mission_service.create_mission(
        db,
        auth,
        ScrapingMissionCreate(
            title="Mission",
            country_code="lb",
            original_prompt="Prompt",
            model_set_id="research-set",
        ),
    )
    assert mission.country_code == "LB"
    assert mission.country_name == "Lebanon"


@pytest.mark.asyncio
async def test_legacy_mission_can_set_country_once_and_locks_after_blueprint(
    db: AsyncSession, auth: AuthContext
):
    await create_model_set(db, auth)
    mission = ScrapingMission(
        org_id=auth.org_id,
        created_by=auth.user.id,
        model_set_id="research-set",
        title="Legacy",
        original_prompt="Prompt",
    )
    db.add(mission)
    await db.flush()
    updated = await mission_service.update_mission(
        db, auth, mission.id, ScrapingMissionUpdate(country_code="JO")
    )
    assert updated.country_code == "JO"
    await create_blueprint_version(
        db,
        auth,
        mission_id=mission.id,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    with pytest.raises(Exception, match="Country cannot be changed"):
        await mission_service.update_mission(
            db, auth, mission.id, ScrapingMissionUpdate(country_code="LB")
        )


async def create_planned_team_plan(db: AsyncSession, auth: AuthContext, monkeypatch) -> ScrapingRun:
    blueprint = await create_active_approved_blueprint(db, auth)
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)
    return await run_service.get_run_row(db, auth, run.id)


@pytest.mark.asyncio
async def test_legacy_mission_without_country_cannot_start_execution(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    await create_model_set(db, auth)
    mission = ScrapingMission(
        org_id=auth.org_id,
        created_by=auth.user.id,
        model_set_id="research-set",
        title="Legacy",
        original_prompt="Prompt",
        status=ScrapingMissionStatus.APPROVED,
    )
    db.add(mission)
    await db.flush()
    blueprint = await create_blueprint_version(
        db,
        auth,
        mission_id=mission.id,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    mission.active_blueprint_id = blueprint.id
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, mission.id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    with pytest.raises(Exception, match="Set a mission country"):
        await execution_service.create_execution(
            db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
        )


@pytest.mark.asyncio
async def test_first_mock_execution_snapshots_country_and_agents(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    assert execution.status == "queued"
    assert execution.country_code == "LB"
    agents = await db.execute(
        select(ScrapingExecutionAgent).where(ScrapingExecutionAgent.execution_id == execution.id)
    )
    assert len(agents.scalars().all()) == len(run.agents)


@pytest.mark.asyncio
async def test_production_execution_rejects_mock_mode(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    with pytest.raises(Exception, match="real source discovery"):
        await execution_service.create_execution(
            db,
            auth,
            run.id,
            ScrapingExecutionCreate(execution_type="initial_full_country", mode="mock"),
        )


@pytest.mark.asyncio
async def test_second_active_execution_for_team_plan_returns_conflict(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    first = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    with pytest.raises(Exception, match="active source discovery execution"):
        await execution_service.create_execution(
            db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
        )
    row = await db.get(ScrapingExecution, first.id)
    assert row is not None
    row.status = ScrapingExecutionStatus.COMPLETED
    await db.flush()
    second = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    assert second.id != first.id


@pytest.mark.asyncio
async def test_source_discovery_worker_persists_candidates_without_facilities(
    db: AsyncSession, auth: AuthContext, monkeypatch, caplog
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    async def fake_discover(service_db, context):
        now = datetime.now(UTC)
        query = ScrapingSourceDiscoveryQuery(
            organization_id=context.organization_id,
            execution_id=context.execution_id,
            coverage_cell_id=context.coverage_cell_id,
            task_id=context.task_id,
            country_code=context.country_code,
            country_name=context.country_name,
            region_code=context.region_code,
            region_name=context.region_name,
            language_code=context.language_code,
            language_name=context.language_name,
            source_category=context.source_category,
            query_text="real rehabilitation registry",
            provider="serper",
            status=SourceDiscoveryQueryStatus.SUCCEEDED,
            requested_at=now,
            completed_at=now,
            result_count=1,
            metadata_json={"purpose": "test"},
        )
        service_db.add(query)
        await service_db.flush()
        service_db.add(
            ScrapingSourceCandidate(
                organization_id=context.organization_id,
                execution_id=context.execution_id,
                coverage_cell_id=context.coverage_cell_id,
                discovery_query_id=query.id,
                provider="serper",
                provider_result_id="https://sante.gouv.fr/source",
                rank=1,
                url="https://sante.gouv.fr/source",
                canonical_url="https://sante.gouv.fr/source",
                domain="sante.gouv.fr",
                title="Real source",
                snippet="Real snippet",
                country_code=context.country_code,
                country_name=context.country_name,
                region_code=context.region_code,
                region_name=context.region_name,
                language_code=context.language_code,
                language_name=context.language_name,
                source_category=context.source_category,
                initial_relevance_score=Decimal("1.0"),
                initial_trust_tier="high",
                status=SourceCandidateStatus.DISCOVERED,
                discovered_at=now,
                metadata_json={"position": 1},
            )
        )
        await service_db.flush()
        return SourceDiscoverySummary(
            provider="serper",
            planned_query_count=1,
            query_count=1,
            succeeded_query_count=1,
            failed_query_count=0,
            candidate_count=1,
            duplicate_candidate_count=0,
            rejected_result_count=0,
        )

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    monkeypatch.setattr(
        "app.services.scraping.execution_orchestrator.source_discovery_service.discover",
        fake_discover,
    )

    async def fail_generate(*args, **kwargs):
        raise AssertionError("MockFacilityGenerator must not run in the real discovery path")

    monkeypatch.setattr(mock_facility_generator, "generate", fail_generate)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    assert execution.status == "queued"
    caplog.set_level("INFO", logger="app.services.scraping.execution_orchestrator")
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    detail = await execution_service.get_detail(db, auth, execution.id)
    assert detail.execution.status == "completed"
    assert detail.execution.started_at is not None
    assert detail.country_profile is not None
    assert detail.coverage_summary_counts
    assert detail.task_summary_counts["completed"] > 0
    assert detail.recent_events[0].sequence_number == 1
    assert detail.execution.mode == "real"
    assert detail.execution.sources_discovered > 0
    assert detail.execution.documents_found > 0
    assert detail.execution.records_extracted == 0
    assert await _execution_facility_count(db, execution.id) == 0
    assert await _execution_source_count(db, execution.id) == 0
    assert await _count_rows(db, RehabilitationFieldEvidence) == 0
    cells = (
        await db.execute(
            select(ScrapingCoverageCell).where(ScrapingCoverageCell.execution_id == execution.id)
        )
    ).scalars().all()
    assert all(cell.status == ScrapingCoverageStatus.PARTIALLY_COVERED for cell in cells)
    retrieval_tasks = (
        await db.execute(
            select(ScrapingTask).where(
                ScrapingTask.execution_id == execution.id,
                ScrapingTask.task_type == "retrieve_source",
            )
        )
    ).scalars().all()
    assert retrieval_tasks
    assert all("source_candidate_id" in task.input_json for task in retrieval_tasks)
    assert all("url" not in task.input_json for task in retrieval_tasks)
    assert await _count_rows(db, ScrapingSourceRetrievalAttempt) == len(retrieval_tasks)
    assert await _count_rows(db, ScrapingSourceDocument) == len(retrieval_tasks)
    events = await db.execute(
        select(ScrapingEvent).where(ScrapingEvent.execution_id == execution.id)
    )
    sequences = [event.sequence_number for event in events.scalars().all()]
    assert sequences == sorted(set(sequences))
    assert "scraping_execution_execution_claim_succeeded" in caplog.text
    assert "scraping_execution_orchestrator_completed" in caplog.text


@pytest.mark.asyncio
async def test_retrieval_task_per_cell_candidate_limit_is_enforced(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_coverage_cell", 2)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_execution", 25)
    mock_source_discovery(monkeypatch, candidate_count=5)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    retrieval_tasks = (
        await db.execute(
            select(ScrapingTask).where(
                ScrapingTask.execution_id == execution.id,
                ScrapingTask.task_type == "retrieve_source",
            )
        )
    ).scalars().all()
    assert len(retrieval_tasks) == 2
    per_cell: dict[str, int] = {}
    for task in retrieval_tasks:
        assert set(task.input_json).issuperset({"source_candidate_id", "idempotency_key", "phase"})
        assert "url" not in task.input_json
        per_cell[task.coverage_cell_id] = per_cell.get(task.coverage_cell_id, 0) + 1
    assert len(per_cell) == 1
    assert max(per_cell.values()) <= 2


@pytest.mark.asyncio
async def test_retrieval_task_per_execution_candidate_limit_is_enforced(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    blueprint.blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["France"],
            "regions": ["Region A", "Region B", "Region C"],
        },
        "languages": ["French"],
        "source_strategy": [
            {"source_type": "official directory", "priority": 1},
        ],
    }
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_coverage_cell", 2)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_execution", 5)
    mock_source_discovery(monkeypatch, candidate_count=5)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    retrieval_tasks = (
        await db.execute(
            select(ScrapingTask)
            .where(
                ScrapingTask.execution_id == execution.id,
                ScrapingTask.task_type == "retrieve_source",
            )
            .order_by(ScrapingTask.priority, ScrapingTask.created_at)
        )
    ).scalars().all()
    assert len(retrieval_tasks) == 5
    per_cell: dict[str, int] = {}
    for task in retrieval_tasks:
        assert set(task.input_json).issuperset({"source_candidate_id", "idempotency_key", "phase"})
        assert "url" not in task.input_json
        per_cell[task.coverage_cell_id] = per_cell.get(task.coverage_cell_id, 0) + 1
    assert len(per_cell) == 3
    assert max(per_cell.values()) <= 2
    assert sorted(per_cell.values()) == [1, 2, 2]


@pytest.mark.asyncio
async def test_retrieval_candidate_selection_is_deterministic_deduped_and_diverse(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_coverage_cell", 3)
    monkeypatch.setattr(get_settings(), "source_retrieval_max_candidates_per_execution", 25)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    agents = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().all()
    await SourceDiscoveryExecutionOrchestrator(db)._ensure_profile_matrix_and_tasks(
        execution, list(agents)
    )
    cell = (
        await db.execute(
            select(ScrapingCoverageCell).where(
                ScrapingCoverageCell.execution_id == execution.id
            )
        )
    ).scalars().first()
    assert cell is not None
    query = ScrapingSourceDiscoveryQuery(
        organization_id=auth.org_id,
        execution_id=execution.id,
        coverage_cell_id=cell.id,
        country_code=execution.country_code,
        country_name=execution.country_name,
        region_code=cell.region_code,
        region_name=cell.region_name,
        language_code=cell.language_code or "en",
        language_name=cell.language_name,
        source_category=cell.source_category,
        query_text="selection query",
        provider="serper",
        status=SourceDiscoveryQueryStatus.SUCCEEDED,
        requested_at=datetime.now(UTC),
        completed_at=datetime.now(UTC),
    )
    db.add(query)
    await db.flush()
    duplicate_query = ScrapingSourceDiscoveryQuery(
        organization_id=auth.org_id,
        execution_id=execution.id,
        coverage_cell_id=cell.id,
        country_code=execution.country_code,
        country_name=execution.country_name,
        region_code=cell.region_code,
        region_name=cell.region_name,
        language_code=cell.language_code or "en",
        language_name=cell.language_name,
        source_category=cell.source_category,
        query_text="selection duplicate query",
        provider="serper",
        status=SourceDiscoveryQueryStatus.SUCCEEDED,
        requested_at=datetime.now(UTC),
        completed_at=datetime.now(UTC),
    )
    db.add(duplicate_query)
    await db.flush()

    rows = [
        ("https://b.example.org/page", "b.example.org", "high", 1, "directory", query.id),
        (
            "https://b.example.org/page",
            "b.example.org",
            "high",
            2,
            "directory duplicate",
            duplicate_query.id,
        ),
        ("https://a.gov.fr/page", "a.gov.fr", "medium", 1, "official registry", query.id),
        ("https://c.example.org/page", "c.example.org", "high", 3, "directory", query.id),
        ("https://d.example.org/page", "d.example.org", "low", 1, "directory", query.id),
    ]
    for index, (url, domain, trust, rank, category, discovery_query_id) in enumerate(rows, start=1):
        db.add(
            ScrapingSourceCandidate(
                organization_id=auth.org_id,
                execution_id=execution.id,
                coverage_cell_id=cell.id,
                discovery_query_id=discovery_query_id,
                provider="serper",
                provider_result_id=f"result-{index}",
                rank=rank,
                url=url,
                canonical_url=url,
                domain=domain,
                title=f"Candidate {index}",
                snippet="Snippet",
                country_code=execution.country_code,
                country_name=execution.country_name,
                region_code=cell.region_code,
                region_name=cell.region_name,
                language_code=cell.language_code or "en",
                language_name=cell.language_name,
                source_category=category,
                initial_relevance_score=Decimal("1.0"),
                initial_trust_tier=trust,
                status=SourceCandidateStatus.DISCOVERED,
                discovered_at=datetime.now(UTC),
            )
        )
    await db.flush()
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)
    first = await orchestrator._select_retrieval_candidates(execution.id, cell.id)
    second = await orchestrator._select_retrieval_candidates(execution.id, cell.id)
    assert [candidate.id for candidate in first] == [candidate.id for candidate in second]
    assert len(first) == 3
    assert len({candidate.canonical_url for candidate in first}) == 3
    assert len({candidate.domain for candidate in first}) == 3
    assert [candidate.domain for candidate in first] == [
        "b.example.org",
        "c.example.org",
        "a.gov.fr",
    ]
    discovery_task = (
        await db.execute(
            select(ScrapingTask).where(
                ScrapingTask.execution_id == execution.id,
                ScrapingTask.coverage_cell_id == cell.id,
                ScrapingTask.task_type == "discover_sources",
            )
        )
    ).scalars().one()
    created = await orchestrator._create_retrieval_tasks(execution, discovery_task)
    assert created == 3
    retrieval_tasks = (
        await db.execute(
            select(ScrapingTask).where(
                ScrapingTask.execution_id == execution.id,
                ScrapingTask.coverage_cell_id == cell.id,
                ScrapingTask.task_type == "retrieve_source",
            )
        )
    ).scalars().all()
    selected_ids = {
        task.input_json["source_candidate_id"]
        for task in retrieval_tasks
    }
    selected_candidates = [
        candidate for candidate in first if candidate.id in selected_ids
    ]
    assert len(selected_candidates) == 3
    assert [
        candidate.canonical_url for candidate in selected_candidates
    ].count("https://b.example.org/page") == 1


@pytest.mark.asyncio
async def test_excel_export_rejects_real_discovery_without_extracted_facilities(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    from app.db.session import get_db

    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)

    app = create_app()

    async def override_db():
        yield db

    async def override_auth():
        return auth

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_auth_context] = override_auth
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.get(f"/api/v1/scraping/executions/{execution.id}/export.xlsx")

    assert response.status_code == 409
    assert "only discovered candidate sources" in response.text
    assert await _execution_facility_count(db, execution.id) == 0
    assert await _execution_source_count(db, execution.id) == 0
    assert await _count_rows(db, RehabilitationFieldEvidence) == 0


@pytest.mark.asyncio
async def test_concurrent_event_creation_uses_unique_monotonic_sequences(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    sessionmaker = async_sessionmaker(db.bind, class_=AsyncSession, expire_on_commit=False)

    async def write_event(index: int) -> None:
        async with sessionmaker() as session:
            await execution_service.emit_event(
                session,
                execution.id,
                "concurrent_event",
                f"Concurrent source-discovery event {index}",
            )
            await session.commit()

    await asyncio.gather(*(write_event(index) for index in range(12)))
    events = (
        await db.execute(
            select(ScrapingEvent)
            .where(ScrapingEvent.execution_id == execution.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    sequences = [event.sequence_number for event in events]
    assert sequences == list(range(1, len(events) + 1))
    assert len(sequences) == len(set(sequences))


@pytest.mark.asyncio
async def test_event_sequences_are_independent_per_execution(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    first = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    first_row = await db.get(ScrapingExecution, first.id)
    assert first_row is not None
    first_row.status = ScrapingExecutionStatus.COMPLETED
    await db.flush()
    second = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await execution_service.emit_event(db, first.id, "first_extra", "First execution extra.")
    await execution_service.emit_event(db, second.id, "second_extra", "Second execution extra.")
    await db.commit()
    first_sequences = (
        await db.execute(
            select(ScrapingEvent.sequence_number)
            .where(ScrapingEvent.execution_id == first.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    second_sequences = (
        await db.execute(
            select(ScrapingEvent.sequence_number)
            .where(ScrapingEvent.execution_id == second.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    assert first_sequences == [1, 2]
    assert second_sequences == [1, 2]


@pytest.mark.asyncio
async def test_source_discovery_worker_rejects_blueprint_without_required_dimensions(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    mission_id = await create_country_mission(db, auth, country_code="BR")
    blueprint = await create_blueprint_version(
        db,
        auth,
        mission_id=mission_id,
        status=ScrapingBlueprintStatus.APPROVED,
        active=True,
    )
    blueprint.blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["Brazil"],
            "regions": [],
        },
        "languages": [],
        "source_strategy": [],
    }
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    detail = await execution_service.get_detail(db, auth, execution.id)
    assert detail.execution.status == "failed"
    assert detail.execution.error_message == "Source discovery execution failed."
    assert detail.country_profile is None
    assert detail.coverage_summary_counts == {}
    assert detail.task_summary_counts == {}
    cells = (
        await db.execute(
            select(ScrapingCoverageCell)
            .where(ScrapingCoverageCell.execution_id == execution.id)
            .order_by(ScrapingCoverageCell.region_name)
        )
    ).scalars().all()
    assert cells == []


@pytest.mark.asyncio
async def test_source_discovery_worker_deduplicates_coverage_dimensions(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint = await create_active_approved_blueprint(db, auth)
    blueprint.blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["Lebanon"],
            "regions": [" Beirut ", "Beirut", " BEIRUT ", "Mount Lebanon"],
        },
        "languages": [" English ", "English"],
        "source_strategy": [
            {"source_type": " official directory ", "priority": 1},
            {"source_type": "Official Directory", "priority": 2},
        ],
    }
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    cells = (
        await db.execute(
            select(ScrapingCoverageCell)
            .where(ScrapingCoverageCell.execution_id == execution.id)
            .order_by(ScrapingCoverageCell.region_name)
        )
    ).scalars().all()
    identities = [
        (cell.region_name, cell.language_name, cell.source_category) for cell in cells
    ]
    assert len(cells) == 2
    assert len(set(identities)) == len(identities)
    assert [cell.region_name for cell in cells] == ["Beirut", "Mount Lebanon"]


def test_coverage_dimensions_keep_region_codes_inside_schema_limit(
    db: AsyncSession,
):
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)
    long_name = (
        "A Very Long Administrative Region Name "
        "That Exceeds The Database Limit"
    )

    regions, languages, categories = (
        orchestrator._coverage_dimensions_from_blueprint(
            {
                "scope": {
                    "regions": [
                        {
                            "code": long_name,
                            "name": long_name,
                        }
                    ]
                },
                "languages": [
                    {
                        "code": " en ",
                        "name": " English ",
                    }
                ],
                "source_strategy": [" general web "],
            },
            "ZZ",
            "Fallback Country",
        )
    )

    assert regions == [
        {
            "code": regions[0]["code"],
            "name": long_name,
        }
    ]
    assert regions[0]["code"]
    assert len(regions[0]["code"]) <= 32
    assert languages == [{"code": "en", "name": "English"}]
    assert categories == ["general web"]


def test_coverage_dimensions_truncate_long_source_categories(db: AsyncSession):
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)
    long_category = (
        "Federal government ministry directories and official national registries "
        "of licensed rehabilitation and addiction treatment facilities " + ("x" * 80)
    )

    _, _, categories = orchestrator._coverage_dimensions_from_blueprint(
        {
            "scope": {"regions": ["Vienna"]},
            "languages": [{"code": "de", "name": "German"}],
            "source_strategy": [{"source_type": long_category}],
        },
        "AT",
        "Austria",
    )

    assert len(categories) == 1
    assert len(categories[0]) <= 120
    assert categories[0] == long_category[:120].rstrip()


def test_coverage_dimensions_derives_codes_for_string_languages(
    db: AsyncSession,
):
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)

    _, languages, _ = orchestrator._coverage_dimensions_from_blueprint(
        {
            "scope": {
                "regions": ["North"],
            },
            "languages": [
                " Spanish ",
                "Catalan",
                "Spanish",
            ],
            "source_strategy": ["directory"],
        },
        "ZZ",
        "Fallback Country",
    )

    assert languages == [
        {"code": "es", "name": "Spanish"},
        {"code": "ca", "name": "Catalan"},
    ]

def test_mock_facility_contexts_rotate_regions_and_languages():
    cells = [
        _CellFallback(
            id=f"{region}-{language}-official",
            region_name=region,
            region_code=region.lower(),
            language_code=language,
            language_name=language.upper(),
            source_category="official",
        )
        for region in ["North", "South", "East"]
        for language in ["es", "ca"]
    ] + [
        _CellFallback(
            id="north-es-directory",
            region_name="North",
            region_code="north",
            language_code="es",
            language_name="ES",
            source_category="directory",
        )
    ]
    contexts = _facility_coverage_contexts(cells, 6)
    assert {context.region_name for context in contexts} == {"North", "South", "East"}
    assert {context.language_code for context in contexts} == {"es", "ca"}
    assert [(context.region_name, context.language_code) for context in contexts] == [
        ("East", "ca"),
        ("North", "es"),
        ("South", "ca"),
        ("East", "es"),
        ("North", "ca"),
        ("South", "es"),
    ]
    retry_contexts = _facility_coverage_contexts(cells, 6)
    assert [context.id for context in retry_contexts] == [context.id for context in contexts]


def test_execution_outcome_labels_completed_with_meaningful_gaps():
    assert execution_outcome_label(ScrapingExecutionStatus.COMPLETED, 0) == "Completed"
    assert execution_outcome_label(ScrapingExecutionStatus.COMPLETED, 1) == "Completed with Gaps"
    assert execution_outcome_label(ScrapingExecutionStatus.FAILED, 5) == "Failed"
    assert execution_outcome_label(ScrapingExecutionStatus.CANCELLED, 5) == "Cancelled"
    assert execution_outcome_label(ScrapingExecutionStatus.QUEUED, 0) == "Queued"
    assert execution_outcome_label(ScrapingExecutionStatus.RUNNING, 0) == "Running"
    assert (
        execution_outcome_label(ScrapingExecutionStatus.CANCEL_REQUESTED, 0)
        == "Cancellation Requested"
    )


@pytest.mark.parametrize(
    "gap_status",
    [
        ScrapingCoverageStatus.PARTIALLY_COVERED,
        ScrapingCoverageStatus.BLOCKED,
        ScrapingCoverageStatus.HUMAN_REVIEW_REQUIRED,
    ],
)
def test_completed_execution_outcome_uses_meaningful_gap_statuses(gap_status):
    coverage = [ScrapingCoverageCell(status=gap_status)]
    assert coverage_gap_count(coverage) == 1
    assert (
        execution_outcome_label(ScrapingExecutionStatus.COMPLETED, coverage_gap_count(coverage))
        == "Completed with Gaps"
    )


def test_coverage_gap_count_excludes_covered_no_results():
    coverage = [
        ScrapingCoverageCell(status=ScrapingCoverageStatus.COVERED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.COVERED_NO_RESULTS),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.PARTIALLY_COVERED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.BLOCKED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.HUMAN_REVIEW_REQUIRED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.FAILED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.CANCELLED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.NOT_STARTED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.QUEUED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.IN_PROGRESS),
    ]
    assert coverage_gap_count(coverage) == 8


def test_completed_execution_outcome_stays_completed_for_covered_cells():
    coverage = [
        ScrapingCoverageCell(status=ScrapingCoverageStatus.COVERED),
        ScrapingCoverageCell(status=ScrapingCoverageStatus.COVERED_NO_RESULTS),
    ]
    assert coverage_gap_count(coverage) == 0
    assert (
        execution_outcome_label(ScrapingExecutionStatus.COMPLETED, coverage_gap_count(coverage))
        == "Completed"
    )


def test_blueprint_judge_prompt_requires_separate_administrative_regions():
    prompt_path = (
        Path(__file__).resolve().parents[1]
        / "app"
        / "prompts"
        / "scraping"
        / "blueprint_judge.j2"
    )
    prompt = prompt_path.read_text(encoding="utf-8")
    assert "one region per array item" in prompt
    assert "first-level administrative subdivision" in prompt
    assert "all provinces and territories" in prompt
    assert "nationwide" in prompt


@pytest.mark.asyncio
async def test_source_discovery_worker_missing_execution_logs_skip_reason(db: AsyncSession, caplog):
    caplog.set_level("INFO", logger="app.services.scraping.execution_orchestrator")
    await SourceDiscoveryExecutionOrchestrator(db).run("missing-execution-id")
    assert logged_execution_reason(caplog, "execution_not_found")


@pytest.mark.asyncio
async def test_source_discovery_worker_terminal_execution_logs_skip_reason(
    db: AsyncSession, auth: AuthContext, monkeypatch, caplog
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    row = await db.get(ScrapingExecution, execution.id)
    assert row is not None
    row.status = ScrapingExecutionStatus.COMPLETED
    await db.flush()
    caplog.set_level("INFO", logger="app.services.scraping.execution_orchestrator")
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    assert logged_execution_reason(caplog, "execution_already_terminal")


@pytest.mark.asyncio
async def test_source_discovery_worker_missing_execution_agents_fails_deterministically(
    db: AsyncSession, auth: AuthContext, monkeypatch, caplog
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    agents = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().all()
    for agent in agents:
        await db.delete(agent)
    await db.flush()
    caplog.set_level("INFO", logger="app.services.scraping.execution_orchestrator")
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    failed = await db.get(ScrapingExecution, execution.id)
    assert failed is not None
    assert failed.status == ScrapingExecutionStatus.FAILED
    assert failed.error_message == "Source discovery execution failed."
    events = await db.execute(
        select(ScrapingEvent).where(
            ScrapingEvent.execution_id == execution.id,
            ScrapingEvent.event_type == "execution_failed",
        )
    )
    assert len(events.scalars().all()) == 1
    assert logged_execution_reason(caplog, "no_execution_agents")


@pytest.mark.asyncio
async def test_source_discovery_worker_accepts_long_descriptive_source_category(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    long_category = (
        "Government Ministry/Registry (Ministry of Public Health, Ministry of Social Affairs)"
    )
    blueprint = await create_active_approved_blueprint(db, auth)
    blueprint.blueprint_json = {
        **valid_blueprint(),
        "source_strategy": [
            {
                "source_type": long_category,
                "priority": 1,
                "trust_tier": "high",
                "purpose": "seed sources",
                "required": True,
            }
        ],
    }
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(execution.id)
    cells = await db.execute(
        select(ScrapingCoverageCell).where(ScrapingCoverageCell.execution_id == execution.id)
    )
    categories = {cell.source_category for cell in cells.scalars().all()}
    assert long_category in categories


@pytest.mark.asyncio
async def test_orchestrator_failure_log_renders_stage_and_exception_type(
    db: AsyncSession, auth: AuthContext, monkeypatch, caplog
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)

    async def fail_after_profile(execution_row, execution_agents):
        orchestrator.current_stage = "flush_coverage_cells"
        orchestrator.coverage_region_count = 2
        orchestrator.coverage_language_count = 1
        orchestrator.coverage_source_category_count = 3
        orchestrator.attempted_coverage_cell_count = 6
        raise RuntimeError("database detail that must not be persisted")

    monkeypatch.setattr(
        orchestrator, "_ensure_profile_matrix_and_tasks", fail_after_profile
    )
    caplog.set_level("ERROR", logger="app.services.scraping.execution_orchestrator")
    await orchestrator.run(execution.id)
    failed = await db.get(ScrapingExecution, execution.id)
    assert failed is not None
    assert failed.status == ScrapingExecutionStatus.FAILED
    assert failed.error_message == "Source discovery execution failed."
    assert "database detail that must not be persisted" not in failed.error_message
    events = await db.execute(
        select(ScrapingEvent).where(
            ScrapingEvent.execution_id == execution.id,
            ScrapingEvent.event_type == "execution_failed",
        )
    )
    failed_events = events.scalars().all()
    assert len(failed_events) == 1
    assert failed_events[0].message == "Source discovery execution failed."
    assert "database detail that must not be persisted" not in failed_events[0].message
    rendered_log = caplog.text
    assert f"scraping_execution_failed execution_id={execution.id}" in rendered_log
    assert "stage=flush_coverage_cells" in rendered_log
    assert "exception_type=RuntimeError" in rendered_log
    assert "region_count=2" in rendered_log
    assert "language_count=1" in rendered_log
    assert "source_category_count=3" in rendered_log
    assert "attempted_coverage_cell_count=6" in rendered_log


@pytest.mark.asyncio
async def test_orchestrator_failed_flush_rolls_back_and_stores_failed_event(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    orchestrator = SourceDiscoveryExecutionOrchestrator(db)

    async def fail_with_pending_rollback(execution_row, execution_agents):
        db.add(
            ScrapingEvent(
                execution_id=execution_row.id,
                sequence_number=1,
                event_type="duplicate_sequence_for_test",
                message="This duplicate event intentionally fails.",
                metadata_json={"phase": "source_discovery"},
            )
        )
        await db.flush()

    monkeypatch.setattr(
        orchestrator, "_ensure_profile_matrix_and_tasks", fail_with_pending_rollback
    )
    await orchestrator.run(execution.id)
    failed = await db.get(ScrapingExecution, execution.id)
    assert failed is not None
    assert failed.status == ScrapingExecutionStatus.FAILED
    assert failed.error_message == "Source discovery execution failed."
    events = await db.execute(
        select(ScrapingEvent).where(
            ScrapingEvent.execution_id == execution.id,
            ScrapingEvent.event_type == "execution_failed",
        )
    )
    assert len(events.scalars().all()) == 1


@pytest.mark.asyncio
async def test_failed_execution_cleanup_terminalizes_active_agents_and_tasks(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    agent = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().first()
    assert agent is not None
    task = ScrapingTask(
        execution_id=execution.id,
        execution_agent_id=agent.id,
        task_type="discover_sources",
        title="Running source discovery task",
        status=ScrapingTaskStatus.RUNNING,
        priority=1,
        input_json={"phase": "source_discovery"},
        output_json={},
        dependency_task_ids_json=[],
        started_at=None,
        current_action="Running",
    )
    db.add(task)
    await db.flush()
    agent.status = ScrapingExecutionAgentStatus.RUNNING
    agent.current_task_id = task.id
    agent.current_action = "Running real source discovery"
    await db.commit()

    await SourceDiscoveryExecutionOrchestrator(db)._mark_failed_safely(execution.id, "RuntimeError")
    failed = await db.get(ScrapingExecution, execution.id)
    assert failed is not None
    assert failed.status == ScrapingExecutionStatus.FAILED
    agents = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().all()
    tasks = (
        await db.execute(
            select(ScrapingTask).where(ScrapingTask.execution_id == execution.id)
        )
    ).scalars().all()
    assert all(agent.status != ScrapingExecutionAgentStatus.RUNNING for agent in agents)
    assert all(agent.status != ScrapingExecutionAgentStatus.WAITING for agent in agents)
    assert all(agent.current_task_id is None for agent in agents)
    assert all(agent.current_action is None for agent in agents)
    assert all(agent.completed_at is not None for agent in agents)
    assert all(task.status == ScrapingTaskStatus.FAILED for task in tasks)
    assert all(task.current_action is None for task in tasks)
    assert all(task.completed_at is not None for task in tasks)
    failed_events = (
        await db.execute(
            select(ScrapingEvent)
            .where(
                ScrapingEvent.execution_id == execution.id,
                ScrapingEvent.event_type == "execution_failed",
            )
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    assert len(failed_events) == 1


def test_scraping_worker_timeout_is_explicit_and_above_default():
    assert WorkerSettings.job_timeout > 300


def test_worker_uses_real_source_discovery_function():
    assert WorkerSettings.functions == [run_scraping_execution]


@pytest.mark.asyncio
async def test_timeout_cleanup_marks_execution_failed_and_terminalizes_children(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    agent = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().first()
    assert agent is not None
    task = ScrapingTask(
        execution_id=execution.id,
        execution_agent_id=agent.id,
        task_type="discover_sources",
        title="Timeout source discovery task",
        status=ScrapingTaskStatus.QUEUED,
        priority=1,
        input_json={"phase": "source_discovery"},
        output_json={},
        dependency_task_ids_json=[],
        current_action="Queued",
    )
    db.add(task)
    agent.status = ScrapingExecutionAgentStatus.RUNNING
    agent.current_action = "Running real source discovery"
    execution.status = ScrapingExecutionStatus.RUNNING
    await db.commit()

    await SourceDiscoveryExecutionOrchestrator(db)._mark_failed_safely(
        execution.id,
        "worker_timeout",
        error_message="Source discovery execution failed after the worker job timed out.",
        event_message="Source discovery execution failed after the worker job timed out.",
    )
    failed = await db.get(ScrapingExecution, execution.id)
    assert failed is not None
    assert failed.status == ScrapingExecutionStatus.FAILED
    assert failed.completed_at is not None
    assert "timed out" in (failed.error_message or "")
    agents = (
        await db.execute(
            select(ScrapingExecutionAgent).where(
                ScrapingExecutionAgent.execution_id == execution.id
            )
        )
    ).scalars().all()
    tasks = (
        await db.execute(
            select(ScrapingTask).where(ScrapingTask.execution_id == execution.id)
        )
    ).scalars().all()
    assert all(agent.status == ScrapingExecutionAgentStatus.FAILED for agent in agents)
    assert all(agent.current_action is None for agent in agents)
    assert all(task.status == ScrapingTaskStatus.FAILED for task in tasks)
    assert all(task.current_action is None for task in tasks)
    events = (
        await db.execute(
            select(ScrapingEvent)
            .where(ScrapingEvent.execution_id == execution.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    assert [event.sequence_number for event in events] == list(range(1, len(events) + 1))
    assert any("timed out" in event.message for event in events)


@pytest.mark.asyncio
async def test_execution_delete_and_team_plan_active_execution_rules(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    execution = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    with pytest.raises(Exception, match="child execution campaign is active"):
        await run_service.delete_run(db, auth, run.id)
    with pytest.raises(Exception, match="Active source discovery executions cannot be deleted"):
        await execution_service.delete_execution(db, auth, execution.id)
    row = await db.get(ScrapingExecution, execution.id)
    assert row is not None
    row.status = ScrapingExecutionStatus.CANCELLED
    await db.flush()
    await execution_service.delete_execution(db, auth, execution.id)
    assert await db.get(ScrapingExecution, execution.id) is None
    agents = await db.execute(
        select(ScrapingExecutionAgent).where(ScrapingExecutionAgent.execution_id == execution.id)
    )
    cells = await db.execute(
        select(ScrapingCoverageCell).where(ScrapingCoverageCell.execution_id == execution.id)
    )
    assert agents.scalars().all() == []
    assert cells.scalars().all() == []


async def run_completed_mock_execution(
    db: AsyncSession, auth: AuthContext, monkeypatch
) -> ScrapingExecution:
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(summary.id)
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    await mock_facility_generator.generate(db, execution)
    await mock_facility_generator.refresh_execution_metrics(db, execution)
    await db.flush()
    return execution


async def run_completed_mock_execution_with_blueprint(
    db: AsyncSession,
    auth: AuthContext,
    monkeypatch,
    blueprint_json: dict,
) -> ScrapingExecution:
    blueprint = await create_active_approved_blueprint(db, auth)
    blueprint.blueprint_json = blueprint_json
    mock_planner(monkeypatch, 3)
    run = await run_service.plan_team(db, auth, blueprint.mission_id)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    mock_source_discovery(monkeypatch)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    await SourceDiscoveryExecutionOrchestrator(db).run(summary.id)
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    await mock_facility_generator.generate(db, execution)
    await mock_facility_generator.refresh_execution_metrics(db, execution)
    await db.flush()
    return execution


def test_mock_attribute_typed_values_dispatch_by_exact_type():
    true_values = _typed_attribute_values(True)
    assert true_values["value_boolean"] is True
    assert true_values["value_number"] is None
    assert true_values["value_text"] is None

    false_values = _typed_attribute_values(False)
    assert false_values["value_boolean"] is False
    assert false_values["value_number"] is None
    assert false_values["value_text"] is None

    integer_values = _typed_attribute_values(18)
    assert integer_values["value_boolean"] is None
    assert integer_values["value_number"] == Decimal("18")
    assert integer_values["value_text"] is None

    float_values = _typed_attribute_values(12.5)
    assert float_values["value_boolean"] is None
    assert float_values["value_number"] == Decimal("12.5")
    assert float_values["value_text"] is None

    decimal_values = _typed_attribute_values(Decimal("1200.25"))
    assert decimal_values["value_boolean"] is None
    assert decimal_values["value_number"] == Decimal("1200.25")
    assert decimal_values["value_text"] is None

    string_values = _typed_attribute_values("Mock available")
    assert string_values["value_boolean"] is None
    assert string_values["value_number"] is None
    assert string_values["value_text"] == "Mock available"

    none_values = _typed_attribute_values(None)
    assert none_values["value_boolean"] is None
    assert none_values["value_number"] is None
    assert none_values["value_text"] is None

    with pytest.raises(TypeError, match="Unsupported mock facility attribute value type"):
        _typed_attribute_values(["unsupported"])


def test_mock_attribute_typed_values_reject_nonfinite_numbers():
    with pytest.raises(ValueError, match="numeric value must be finite"):
        _typed_attribute_values(float("nan"))
    with pytest.raises(ValueError, match="numeric value must be finite"):
        _typed_attribute_values(float("inf"))
    with pytest.raises(ValueError, match="Decimal value must be finite"):
        _typed_attribute_values(Decimal("NaN"))


@pytest.mark.asyncio
async def test_mock_execution_persists_idempotent_rehabilitation_dataset(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    execution = await run_completed_mock_execution(db, auth, monkeypatch)
    facilities = (
        await db.execute(
            select(RehabilitationFacility)
            .options(selectinload(RehabilitationFacility.locations))
            .where(RehabilitationFacility.execution_id == execution.id)
            .order_by(RehabilitationFacility.stable_key)
        )
    ).scalars().all()
    assert 4 <= len(facilities) <= 8
    assert all(facility.is_mock for facility in facilities)
    assert all("Mock" in facility.canonical_name for facility in facilities)
    assert len({facility.stable_key for facility in facilities}) == len(facilities)
    sources = (
        await db.execute(select(RehabilitationSource).where(RehabilitationSource.execution_id == execution.id))
    ).scalars().all()
    assert sources
    assert all(source.is_mock for source in sources)
    assert {source.domain for source in sources} == {"example.invalid"}
    attributes = (
        await db.execute(
            select(RehabilitationFacilityAttribute).where(
                RehabilitationFacilityAttribute.facility_id.in_(
                    [facility.id for facility in facilities]
                )
            )
        )
    ).scalars().all()
    assert attributes
    boolean_attribute = next(
        attribute for attribute in attributes if attribute.attribute_key == "counseling"
    )
    assert boolean_attribute.value_boolean is True
    assert boolean_attribute.value_number is None
    assert boolean_attribute.value_text is None
    false_attribute = next(
        attribute
        for attribute in attributes
        if attribute.attribute_key == "detoxification" and attribute.value_boolean is False
    )
    assert false_attribute.value_boolean is False
    assert false_attribute.value_number is None
    numeric_attribute = next(
        attribute for attribute in attributes if attribute.attribute_key == "minimum_age"
    )
    assert Decimal(str(numeric_attribute.value_number)) == Decimal("18.00")
    assert numeric_attribute.value_boolean is None
    text_attribute = next(
        attribute for attribute in attributes if attribute.attribute_key == "residential_30_day"
    )
    assert text_attribute.value_text == "Mock available"
    assert text_attribute.value_boolean is None
    assert text_attribute.value_number is None
    staff_names = (
        await db.execute(
            select(func.count()).select_from(RehabilitationFacility).where(
                RehabilitationFacility.execution_id == execution.id
            )
        )
    ).scalar_one()
    assert staff_names == len(facilities)
    first_counts = {
        "facilities": len(facilities),
        "contacts": await _count_rows(db, RehabilitationFacilityContact),
        "sources": len(sources),
        "evidence": await _count_rows(db, RehabilitationFieldEvidence),
        "unresolved": await _count_rows(db, RehabilitationUnresolvedField),
    }
    await mock_facility_generator.generate(db, execution)
    await mock_facility_generator.refresh_execution_metrics(db, execution)
    await db.flush()
    second_counts = {
        "facilities": await _execution_facility_count(db, execution.id),
        "contacts": await _count_rows(db, RehabilitationFacilityContact),
        "sources": await _execution_source_count(db, execution.id),
        "evidence": await _count_rows(db, RehabilitationFieldEvidence),
        "unresolved": await _count_rows(db, RehabilitationUnresolvedField),
    }
    assert second_counts == first_counts
    assert execution.records_extracted == first_counts["facilities"]
    assert execution.sources_discovered == first_counts["sources"]
    assert execution.records_verified == len(
        [facility for facility in facilities if facility.verification_status == "verified"]
    )
    assert execution.duplicates_detected == 1


@pytest.mark.asyncio
async def test_rehabilitation_dataset_relationships_and_statuses(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    execution = await run_completed_mock_execution(db, auth, monkeypatch)
    facilities = (
        await db.execute(
            select(RehabilitationFacility)
            .options(selectinload(RehabilitationFacility.locations))
            .where(RehabilitationFacility.execution_id == execution.id)
            .order_by(RehabilitationFacility.stable_key)
        )
    ).scalars().all()
    evidence = (
        await db.execute(select(RehabilitationFieldEvidence).order_by(RehabilitationFieldEvidence.created_at))
    ).scalars().all()
    assert evidence
    assert all(row.facility_id in {facility.id for facility in facilities} for row in evidence)
    assert all(row.source_id is not None for row in evidence)
    shared_sources = (
        await db.execute(
            select(RehabilitationSource)
            .where(RehabilitationSource.execution_id == execution.id)
            .order_by(RehabilitationSource.canonical_url)
        )
    ).scalars().all()
    assert any(source.canonical_url.endswith("/mock-index") for source in shared_sources)
    duplicate = (
        await db.execute(
            select(RehabilitationPossibleDuplicate).where(
                RehabilitationPossibleDuplicate.execution_id == execution.id
            )
        )
    ).scalar_one()
    assert duplicate.left_facility_id < duplicate.right_facility_id
    unresolved_statuses = {
        status
        for (status,) in (
            await db.execute(select(RehabilitationUnresolvedField.unresolved_status))
        ).all()
    }
    assert {"searched_not_found", "conflicting"}.issubset(unresolved_statuses)


@pytest.mark.asyncio
async def test_mock_facilities_distribute_across_persisted_regions_and_languages(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["Mock Country"],
            "regions": ["Mock North", "Mock South", "Mock East"],
        },
        "languages": ["Spanish", "Catalan"],
        "source_strategy": [
            {
                "source_type": "official directory",
                "priority": 1,
                "trust_tier": "high",
                "purpose": "seed sources",
                "required": True,
            },
            {
                "source_type": "professional registry",
                "priority": 2,
                "trust_tier": "medium",
                "purpose": "supporting sources",
                "required": False,
            },
        ],
    }
    execution = await run_completed_mock_execution_with_blueprint(
        db, auth, monkeypatch, blueprint_json
    )
    facilities = (
        await db.execute(
            select(RehabilitationFacility)
            .options(selectinload(RehabilitationFacility.locations))
            .where(RehabilitationFacility.execution_id == execution.id)
            .order_by(RehabilitationFacility.stable_key)
        )
    ).scalars().all()
    cells = (
        await db.execute(
            select(ScrapingCoverageCell).where(ScrapingCoverageCell.execution_id == execution.id)
        )
    ).scalars().all()
    cell_regions = {cell.region_name for cell in cells}
    cell_languages = {cell.language_code for cell in cells}
    assert {"es", "ca"}.issubset(cell_languages)
    assert "" not in cell_languages
    facility_regions = {facility.primary_region for facility in facilities}
    assert len(facility_regions) > 1
    assert facility_regions.issubset(cell_regions)
    assert all(facility.latitude is None and facility.longitude is None for facility in facilities)

    locations = [
        location
        for facility in facilities
        for location in facility.locations
        if location.is_primary
    ]
    assert all(location.region in cell_regions for location in locations)
    assert all(
        location.region
        == next(facility.primary_region for facility in facilities if facility.id == location.facility_id)
        for location in locations
    )

    sources = (
        await db.execute(
            select(RehabilitationSource).where(RehabilitationSource.execution_id == execution.id)
        )
    ).scalars().all()
    source_language_codes = {source.language_code for source in sources}
    assert "" not in source_language_codes
    assert source_language_codes.issubset(cell_languages)
    assert len(source_language_codes) > 1
    assert {source.region for source in sources}.issubset(cell_regions)
    source_ids = [source.id for source in sources]
    evidence_rows = (
        await db.execute(
            select(RehabilitationFieldEvidence).where(
                RehabilitationFieldEvidence.source_id.in_(source_ids)
            )
        )
    ).scalars().all()
    source_languages = {source.id: source.language_code for source in sources}
    assert all(
        evidence.language_code == source_languages[evidence.source_id]
        for evidence in evidence_rows
        if evidence.source_id
    )
    facility_snapshot = [
        (facility.stable_key, facility.primary_region, facility.primary_city)
        for facility in facilities
    ]
    source_snapshot = [
        (source.canonical_url, source.region, source.language_code, source.source_category)
        for source in sorted(sources, key=lambda source: source.canonical_url)
    ]
    evidence_snapshot = [
        (evidence.field_path, evidence.extracted_value, evidence.language_code)
        for evidence in sorted(
            evidence_rows,
            key=lambda evidence: (
                evidence.field_path,
                evidence.extracted_value or "",
                evidence.language_code or "",
            ),
        )
    ]
    child_counts = {
        "facilities": await _execution_facility_count(db, execution.id),
        "sources": await _execution_source_count(db, execution.id),
        "evidence": len(evidence_rows),
        "attributes": await _count_rows(db, RehabilitationFacilityAttribute),
        "contacts": await _count_rows(db, RehabilitationFacilityContact),
    }
    await mock_facility_generator.generate(db, execution)
    await db.flush()
    retry_facilities = (
        await db.execute(
            select(RehabilitationFacility)
            .where(RehabilitationFacility.execution_id == execution.id)
            .order_by(RehabilitationFacility.stable_key)
        )
    ).scalars().all()
    retry_sources = (
        await db.execute(
            select(RehabilitationSource).where(RehabilitationSource.execution_id == execution.id)
        )
    ).scalars().all()
    retry_source_ids = [source.id for source in retry_sources]
    retry_evidence_rows = (
        await db.execute(
            select(RehabilitationFieldEvidence).where(
                RehabilitationFieldEvidence.source_id.in_(retry_source_ids)
            )
        )
    ).scalars().all()
    assert [
        (facility.stable_key, facility.primary_region, facility.primary_city)
        for facility in retry_facilities
    ] == facility_snapshot
    assert [
        (source.canonical_url, source.region, source.language_code, source.source_category)
        for source in sorted(retry_sources, key=lambda source: source.canonical_url)
    ] == source_snapshot
    assert [
        (evidence.field_path, evidence.extracted_value, evidence.language_code)
        for evidence in sorted(
            retry_evidence_rows,
            key=lambda evidence: (
                evidence.field_path,
                evidence.extracted_value or "",
                evidence.language_code or "",
            ),
        )
    ] == evidence_snapshot
    assert {
        "facilities": await _execution_facility_count(db, execution.id),
        "sources": await _execution_source_count(db, execution.id),
        "evidence": len(retry_evidence_rows),
        "attributes": await _count_rows(db, RehabilitationFacilityAttribute),
        "contacts": await _count_rows(db, RehabilitationFacilityContact),
    } == child_counts


@pytest.mark.asyncio
async def test_multi_dimension_mock_execution_emits_unique_ordered_events(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["Mock Country"],
            "regions": [f"Mock Region {index}" for index in range(1, 6)],
        },
        "languages": ["English", "Spanish", "Catalan"],
        "source_strategy": [
            {"source_type": f"mock source {index}", "priority": index}
            for index in range(1, 4)
        ],
    }
    execution = await run_completed_mock_execution_with_blueprint(
        db, auth, monkeypatch, blueprint_json
    )
    events = (
        await db.execute(
            select(ScrapingEvent)
            .where(ScrapingEvent.execution_id == execution.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    sequences = [event.sequence_number for event in events]
    assert len(events) > 50
    assert sequences == list(range(1, len(events) + 1))
    assert len(sequences) == len(set(sequences))


@pytest.mark.asyncio
async def test_large_mock_execution_batches_metric_refreshes(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    refresh_calls = 0
    original_refresh = SourceDiscoveryExecutionOrchestrator._refresh_metrics

    async def counted_refresh(self, execution):
        nonlocal refresh_calls
        refresh_calls += 1
        await original_refresh(self, execution)

    monkeypatch.setattr(SourceDiscoveryExecutionOrchestrator, "_refresh_metrics", counted_refresh)
    blueprint_json = {
        **valid_blueprint(),
        "scope": {
            "included": ["public listings"],
            "excluded": ["private data"],
            "countries": ["Mock Country"],
            "regions": [f"Mock Region {index}" for index in range(1, 20)],
        },
        "languages": ["English", "Spanish", "Catalan", "French"],
        "source_strategy": [
            {"source_type": f"mock source {index}", "priority": index}
            for index in range(1, 7)
        ],
    }
    execution = await run_completed_mock_execution_with_blueprint(
        db, auth, monkeypatch, blueprint_json
    )
    task_count = int(
        (
            await db.execute(
                select(func.count(ScrapingTask.id)).where(
                    ScrapingTask.execution_id == execution.id,
                    ScrapingTask.task_type == "discover_sources",
                )
            )
        ).scalar_one()
    )
    assert task_count == 19 * 4 * 6
    assert 1 < refresh_calls <= (task_count // METRIC_REFRESH_TASK_INTERVAL) + 3
    facility_count = await _execution_facility_count(db, execution.id)
    source_count = await _execution_source_count(db, execution.id)
    assert execution.status == ScrapingExecutionStatus.COMPLETED
    assert execution.records_extracted == facility_count
    assert execution.sources_discovered == source_count
    events = (
        await db.execute(
            select(ScrapingEvent.sequence_number)
            .where(ScrapingEvent.execution_id == execution.id)
            .order_by(ScrapingEvent.sequence_number)
        )
    ).scalars().all()
    assert events == list(range(1, len(events) + 1))


@pytest.mark.asyncio
async def test_cancel_requested_execution_does_not_generate_rehabilitation_dataset(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    run = await create_planned_team_plan(db, auth, monkeypatch)

    async def fake_enqueue(execution_id):
        return None

    monkeypatch.setattr(execution_service, "enqueue_execution", fake_enqueue)
    summary = await execution_service.create_execution(
        db, auth, run.id, ScrapingExecutionCreate(execution_type="initial_full_country")
    )
    execution = await db.get(ScrapingExecution, summary.id)
    assert execution is not None
    execution.status = ScrapingExecutionStatus.CANCEL_REQUESTED
    await db.flush()
    await mock_facility_generator.generate(db, execution)
    assert await _execution_facility_count(db, execution.id) == 0


@pytest.mark.asyncio
async def test_authorized_facility_api_lists_deterministic_summaries(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    execution = await run_completed_mock_execution(db, auth, monkeypatch)
    rows = await execution_service.list_facilities(db, auth, execution.id)
    assert rows
    assert [row.stable_key for row in rows] == sorted(row.stable_key for row in rows)
    assert rows[0].is_mock is True
    assert rows[0].verification_status in {"verified", "unverified"}
    assert rows[0].confidence_score > 0
    assert rows[0].source_count > 0
    other = await create_other_auth(db)
    with pytest.raises(Exception, match="ScrapingExecution not found"):
        await execution_service.list_facilities(db, other, execution.id)


@pytest.mark.asyncio
async def test_completed_execution_without_coverage_gaps_keeps_completed_label(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    execution = await run_completed_mock_execution(db, auth, monkeypatch)
    await _force_coverage_statuses(
        db,
        execution,
        [ScrapingCoverageStatus.COVERED, ScrapingCoverageStatus.COVERED_NO_RESULTS],
    )
    detail = await execution_service.get_detail(db, auth, execution.id)
    assert detail.execution.status == "completed"
    assert detail.execution.status_label == "Completed"
    assert detail.execution.coverage_debt == 0
    row = await db.get(ScrapingExecution, execution.id)
    assert row is not None
    assert row.status == ScrapingExecutionStatus.COMPLETED


@pytest.mark.asyncio
async def test_excel_export_workbook_contract_and_active_rejection(
    db: AsyncSession, auth: AuthContext, monkeypatch
):
    from app.db.session import get_db
    from openpyxl import load_workbook

    execution = await run_completed_mock_execution(db, auth, monkeypatch)
    await _force_coverage_statuses(
        db,
        execution,
        [ScrapingCoverageStatus.PARTIALLY_COVERED],
    )
    app = create_app()

    async def override_db():
        yield db

    async def override_auth():
        return auth

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_auth_context] = override_auth
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        detail_response = await client.get(f"/api/v1/scraping/executions/{execution.id}")
        response = await client.get(f"/api/v1/scraping/executions/{execution.id}/export.xlsx")
    assert detail_response.status_code == 200
    detail_body = detail_response.json()["execution"]
    assert detail_body["status"] == "completed"
    assert detail_body["status_label"] == "Completed with Gaps"
    assert detail_body["coverage_debt"] > 0
    assert execution.status == ScrapingExecutionStatus.COMPLETED
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "mock-rehabilitation-dataset" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == SHEET_ORDER
    rehab_sheet = workbook["Rehab Centers"]
    assert [rehab_sheet.cell(row=1, column=column).value for column in range(1, 6)] == [
        "Official Name",
        "Alternative Names",
        "Original-Language Name",
        "Facility Type",
        "Organization Type",
    ]
    rehab_headers = [cell.value for cell in rehab_sheet[1]]
    assert "Facility ID" in rehab_headers
    assert rehab_sheet.max_row == execution.records_extracted + 1
    confidence_column = rehab_headers.index("Confidence Score") + 1
    mock_column = rehab_headers.index("Mock") + 1
    website_column = rehab_headers.index("Primary Website") + 1
    latitude_column = rehab_headers.index("Latitude") + 1
    longitude_column = rehab_headers.index("Longitude") + 1
    status_column = rehab_headers.index("Duplicate Status") + 1
    assert isinstance(rehab_sheet.cell(row=2, column=confidence_column).value, int | float)
    assert rehab_sheet.cell(row=2, column=confidence_column).number_format == "0.0%"
    assert rehab_sheet.cell(row=2, column=mock_column).value == "Yes"
    assert rehab_sheet.cell(row=2, column=website_column).hyperlink is not None
    assert rehab_sheet.cell(row=2, column=status_column).value in {
        "Unique",
        "Possible Duplicate",
    }
    assert rehab_sheet.cell(row=2, column=latitude_column).value is None
    assert rehab_sheet.cell(row=2, column=longitude_column).value is None
    assert workbook["Contacts"].max_row - 1 == await _normal_contact_count(db)
    contact_headers = [cell.value for cell in workbook["Contacts"][1]]
    contact_value_column = contact_headers.index("Value") + 1
    assert workbook["Contacts"].cell(row=2, column=contact_value_column).hyperlink is None
    assert workbook["Social Media"].max_row - 1 == await _social_contact_count(db)
    assert workbook["Social Media"].cell(row=2, column=6).hyperlink is not None
    assert workbook["Sources"].max_row - 1 == execution.sources_discovered
    assert workbook["Sources"].cell(row=2, column=6).hyperlink is not None
    assert workbook["Field Evidence"].max_row - 1 == await _count_rows(db, RehabilitationFieldEvidence)
    assert workbook["Possible Duplicates"].max_row - 1 == execution.duplicates_detected
    assert workbook["Unresolved Records"].max_row - 1 == await _count_rows(db, RehabilitationUnresolvedField)
    coverage_sheet = workbook["Coverage Report"]
    assert coverage_sheet["A1"].value == "Coverage Status"
    assert coverage_sheet["A2"].value == "Total Coverage Cells"
    assert coverage_sheet["B2"].value == await _count_rows(db, ScrapingCoverageCell)
    assert coverage_sheet["A14"].value == "Coverage Cell ID"
    assert coverage_sheet.max_row - 14 == await _count_rows(db, ScrapingCoverageCell)
    assert workbook["Execution Summary"]["A1"].value.startswith("MOCK REHABILITATION DATASET")
    assert workbook["Execution Summary"]["A2"].value == (
        "All facility records in this workbook were generated for testing."
    )
    assert workbook["Execution Summary"]["A4"].value == (
        "The facility rows in this workbook are fictional sample records used to test "
        "the dataset structure. They are not a count or estimate of real rehabilitation "
        "centers in the selected country."
    )
    assert workbook["Execution Summary"]["A5"].value == "KPI"
    assert workbook["Execution Summary"]["A6"].value == "Total Facilities"
    assert workbook["Execution Summary"]["B6"].value == execution.records_extracted
    assert workbook["Execution Summary"]["C6"].value == "Sample Facility Count"
    assert workbook["Execution Summary"]["D6"].value == execution.records_extracted
    assert workbook["Execution Summary"]["A10"].value == "Coverage Percentage"
    assert workbook["Execution Summary"]["B10"].number_format == "0.0%"
    assert workbook["Execution Summary"]["C10"].value == "Coverage Outcome"
    assert workbook["Execution Summary"]["D10"].value == "Completed with Gaps"
    summary_values = {
        workbook["Execution Summary"].cell(row=row, column=1).value:
        workbook["Execution Summary"].cell(row=row, column=2).value
        for row in range(1, workbook["Execution Summary"].max_row + 1)
    }
    assert summary_values["Dataset Type"] == "Mock Sample Dataset"
    assert summary_values["Sample Facility Count"] == execution.records_extracted
    assert summary_values["Country Completeness"] == "Not measured in mock mode"
    assert summary_values["Coverage Outcome"] == "Completed with Gaps"
    assert all(workbook[sheet].max_row >= 1 for sheet in SHEET_ORDER)

    queued = await execution_service.create_execution(
        db,
        auth,
        execution.team_plan_id,
        ScrapingExecutionCreate(execution_type="initial_full_country"),
    )
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        active_response = await client.get(f"/api/v1/scraping/executions/{queued.id}/export.xlsx")
    assert active_response.status_code == 409


def test_excel_safe_cell_escapes_formula_like_text():
    assert safe_cell("=SUM(1,1)") == "'=SUM(1,1)"
    assert safe_cell("+CMD") == "'+CMD"
    assert safe_cell("-1+2") == "'-1+2"
    assert safe_cell("@IMPORT") == "'@IMPORT"
    assert safe_cell(12) == 12


async def _count_rows(db: AsyncSession, model) -> int:
    return int((await db.execute(select(func.count(model.id)))).scalar_one())


async def _execution_facility_count(db: AsyncSession, execution_id: str) -> int:
    return int(
        (
            await db.execute(
                select(func.count(RehabilitationFacility.id)).where(
                    RehabilitationFacility.execution_id == execution_id
                )
            )
        ).scalar_one()
    )


async def _execution_source_count(db: AsyncSession, execution_id: str) -> int:
    return int(
        (
            await db.execute(
                select(func.count(RehabilitationSource.id)).where(
                    RehabilitationSource.execution_id == execution_id
                )
            )
        ).scalar_one()
    )


async def _force_coverage_statuses(
    db: AsyncSession,
    execution: ScrapingExecution,
    statuses: list[ScrapingCoverageStatus],
) -> None:
    cells = (
        await db.execute(
            select(ScrapingCoverageCell)
            .where(ScrapingCoverageCell.execution_id == execution.id)
            .order_by(ScrapingCoverageCell.created_at)
        )
    ).scalars().all()
    assert cells
    for index, cell in enumerate(cells):
        cell.status = statuses[index] if index < len(statuses) else ScrapingCoverageStatus.COVERED
    execution.coverage_debt = coverage_gap_count(list(cells))
    await db.flush()
    await db.refresh(execution)


async def _normal_contact_count(db: AsyncSession) -> int:
    return int(
        (
            await db.execute(
                select(func.count(RehabilitationFacilityContact.id)).where(
                    RehabilitationFacilityContact.contact_type.not_in(
                        ["facebook", "instagram", "linkedin", "youtube", "other_social"]
                    )
                )
            )
        ).scalar_one()
    )


async def _social_contact_count(db: AsyncSession) -> int:
    return int(
        (
            await db.execute(
                select(func.count(RehabilitationFacilityContact.id)).where(
                    RehabilitationFacilityContact.contact_type.in_(
                        ["facebook", "instagram", "linkedin", "youtube", "other_social"]
                    )
                )
            )
        ).scalar_one()
    )
