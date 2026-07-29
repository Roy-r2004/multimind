"""Excel workbook export for persisted rehabilitation execution datasets."""

from __future__ import annotations

import re
from collections.abc import Iterable
from datetime import UTC, date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.dependencies import AuthContext
from app.core.exceptions import ConflictError, NotFoundError
from app.db.models import (
    RehabilitationFacility,
    ScrapingExecution,
    ScrapingExecutionStatus,
)
from app.services.scraping.result_metrics import (
    normalized_publication_class,
    primary_phone_for_facility,
)

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TERMINAL_STATUSES = {
    ScrapingExecutionStatus.COMPLETED,
    ScrapingExecutionStatus.FAILED,
    ScrapingExecutionStatus.CANCELLED,
}
SHEET_ORDER = ["Facilities"]
URL_HEADERS = {"Website"}


class ExecutionExportService:
    async def build_workbook(
        self, db: AsyncSession, auth: AuthContext, execution_id: str
    ) -> tuple[bytes, str]:
        data = await self._load(db, auth, execution_id)
        execution = data["execution"]
        if execution.status not in TERMINAL_STATUSES:
            raise ConflictError("Excel report available after execution finishes.")

        workbook = Workbook()
        workbook.properties.title = "Scraping Execution Export"
        workbook.properties.creator = "MultiAI Verdict"
        workbook.remove(workbook.active)
        for sheet_name in SHEET_ORDER:
            workbook.create_sheet(sheet_name)

        self._write_facilities(workbook["Facilities"], data["facilities"])
        for worksheet in workbook.worksheets:
            self._style_sheet(worksheet)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue(), _filename(execution)

    async def _load(self, db: AsyncSession, auth: AuthContext, execution_id: str) -> dict[str, Any]:
        result = await db.execute(
            select(ScrapingExecution).where(
                ScrapingExecution.id == execution_id,
                ScrapingExecution.organization_id == auth.org_id,
            )
        )
        execution = result.scalar_one_or_none()
        if execution is None:
            raise NotFoundError("ScrapingExecution", execution_id)

        facilities = (
            await db.execute(
                select(RehabilitationFacility)
                .where(RehabilitationFacility.execution_id == execution.id)
                .options(
                    selectinload(RehabilitationFacility.contacts),
                )
                .order_by(RehabilitationFacility.canonical_name, RehabilitationFacility.stable_key)
            )
        ).scalars().all()
        return {
            "execution": execution,
            "facilities": list(facilities),
        }

    def _write_facilities(self, ws: Any, facilities: list[RehabilitationFacility]) -> None:
        headers = ["Facility", "Contact", "Website"]
        rows = []
        for facility in facilities:
            if normalized_publication_class(getattr(facility, "publication_class", None)) == "excluded":
                continue
            rows.append(
                [
                    facility.canonical_name,
                    _facility_contact(facility),
                    facility.primary_website,
                ]
            )
        if not rows:
            rows.append(["No facilities recorded."] + [None] * (len(headers) - 1))
        _write_table(ws, headers, rows)

    def _style_sheet(self, ws: Any) -> None:
        header_fill = PatternFill("solid", fgColor="244A6B")
        stripe_fill = PatternFill("solid", fgColor="F8FAFC")
        border = Border(bottom=Side(style="thin", color="D7DEE8"))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
        for row in ws.iter_rows():
            for cell in row:
                header = _column_header(ws, cell.column)
                cell.alignment = Alignment(wrap_text=False, vertical="top")
                cell.border = border
                if cell.row > 1 and cell.row % 2 == 0:
                    cell.fill = stripe_fill
                if header in URL_HEADERS and isinstance(cell.value, str) and _safe_http_url(cell.value):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        for column_index in range(1, ws.max_column + 1):
            letter = get_column_letter(column_index)
            ws.column_dimensions[letter].width = _column_width(ws, column_index)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"


def _write_table(ws: Any, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([safe_cell(value) for value in row])


def safe_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        dt = value if value.tzinfo else value.replace(tzinfo=UTC)
        return dt.astimezone(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="minutes")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    return value


def _facility_contact(facility: RehabilitationFacility) -> str | None:
    primary = (getattr(facility, "primary_contact", None) or "").strip()
    if primary:
        return primary
    phone = primary_phone_for_facility(facility)
    if phone:
        return phone
    contacts = list(getattr(facility, "contacts", None) or [])
    for contact in contacts:
        if getattr(contact, "is_primary", False) and getattr(contact, "value", None):
            return contact.value
    for contact in contacts:
        value = getattr(contact, "value", None)
        if value:
            return value
    return None


def _column_header(ws: Any, column_index: int) -> str | None:
    value = ws.cell(row=1, column=column_index).value
    return str(value) if value is not None else None


def _column_width(ws: Any, column_index: int) -> int:
    header = _column_header(ws, column_index) or ""
    if header == "Facility":
        return 36
    if header == "Website":
        return 42
    if header == "Contact":
        return 28
    return min(max(len(header) + 2, 14), 28)


def _safe_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _filename(execution: ScrapingExecution) -> str:
    country = re.sub(r"[^A-Za-z0-9]+", "-", execution.country_name).strip("-").lower() or "country"
    suffix = execution.id[:8]
    return f"scraping-execution-{country}-{suffix}.xlsx"


execution_export_service = ExecutionExportService()
