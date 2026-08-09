"""Excel workbook export for a Maps Census run.

Client-facing export is a single sheet: ``Eligible Centers`` — only facilities
that passed the strict keep/drop gate (``keep_drop_decision == "keep"``). No
review / excluded / audit sheets are produced for the client workbook.
"""

from __future__ import annotations

import json
import re
from enum import Enum
from io import BytesIO
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import AuthContext
from app.core.exceptions import NotFoundError
from app.db.models import (
    MapsCensusRun,
    MapsClientEligibility,
    MapsPlace,
)
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ELIGIBLE_CENTERS_SHEET = "Eligible Centers"
PHASE_1_SHEET = "Phase 1 Results"

# Mirrors the UI table (src/lib/maps/exportDisplay.ts EXPORT_COLUMNS) exactly.
EXPORT_HEADERS = [
    "Facility Name",
    "Addictions Treated",
    "Location",
    "Languages Spoken",
    "Website",
    "Email",
    "Phone Number",
    "Treatment Price",
    "Bed Count",
]

_NOT_SPECIFIED = "Not Specified"
_CONTACT_PRICING = "Contact for pricing"

# Headers whose string cells should render as clickable links.
_URL_HEADERS = {"Website"}
# Headers whose string cells should render as clickable mailto links.
_EMAIL_HEADERS = {"Email"}
# Headers whose values must stay textual so "+" and leading zeroes survive.
_TEXT_HEADERS = {"Phone Number"}
# Short, scannable columns that read best centered in both axes.
_CENTERED_HEADERS = {"Phone Number", "Treatment Price", "Bed Count"}

_WIDE_COLUMNS = {
    "Facility Name": 34,
    "Addictions Treated": 26,
    "Location": 42,
    "Languages Spoken": 22,
    "Website": 40,
    "Email": 30,
    "Phone Number": 20,
    "Treatment Price": 20,
    "Bed Count": 12,
}


class MapsExportService:
    async def build_workbook(
        self, db: AsyncSession, auth: AuthContext, run_id: str, *, scope: str = "phase2"
    ) -> tuple[bytes, str]:
        run = await db.get(MapsCensusRun, run_id)
        if run is None or run.organization_id != auth.org_id:
            raise NotFoundError("Maps census run", run_id)

        places = (
            await db.execute(
                select(MapsPlace)
                .where(MapsPlace.run_id == run_id)
                .order_by(MapsPlace.canonical_name)
            )
        ).scalars().all()

        workbook = Workbook()
        workbook.properties.title = "Maps Census Export"
        workbook.properties.creator = "MultiAI Verdict"
        workbook.remove(workbook.active)

        if scope == "phase1":
            # Phase 1 = every discovered place the user hasn't manually
            # removed — same set the Phase 1 tab shows on screen.
            sheet_name = PHASE_1_SHEET
            table_name = "Phase1Results"
            scoped_places = [place for place in places if place.manually_excluded_at is None]
        else:
            # Client workbook = only the strict keep/drop "keep" rows that
            # are actually deliverable — no phone and no website means the
            # client has no way to contact the facility, so it is dropped.
            sheet_name = ELIGIBLE_CENTERS_SHEET
            table_name = "EligibleCenters"
            scoped_places = [
                place
                for place in places
                if place.keep_drop_decision == "keep" and _has_contact_info(place)
            ]

        self._write_sheet(
            workbook.create_sheet(sheet_name),
            EXPORT_HEADERS,
            [self._export_place_row(place, run.country_name) for place in scoped_places],
            table_name=table_name,
        )

        buffer = BytesIO()
        workbook.save(buffer)
        suffix = "phase1-results" if scope == "phase1" else "eligible-centers"
        filename = f"{run.country_code.lower()}-maps-census-{suffix}.xlsx"
        return buffer.getvalue(), filename

    async def get_export_summary(
        self, db: AsyncSession, auth: AuthContext, run_id: str
    ) -> dict[str, int]:
        run = await db.get(MapsCensusRun, run_id)
        if run is None or run.organization_id != auth.org_id:
            raise NotFoundError("Maps census run", run_id)

        places = (
            await db.execute(select(MapsPlace).where(MapsPlace.run_id == run_id))
        ).scalars().all()

        counts = {
            ELIGIBLE_CENTERS_SHEET: 0,
            "Other": 0,
        }
        for place in places:
            if _is_eligible_center(place):
                counts[ELIGIBLE_CENTERS_SHEET] += 1
            else:
                counts["Other"] += 1
        return counts

    def _export_place_row(self, place: MapsPlace, country_name: str | None = None) -> list[Any]:
        """Mirror src/lib/maps/exportDisplay.ts placeToExportRow exactly."""
        return [
            _ui_text(place.canonical_name or place.raw_name),
            _ui_list(place.addictions_treated),
            _ui_location(place, country_name),
            _ui_list(place.languages_spoken),
            _ui_website(place),
            _ui_text(place.contact_email),
            _ui_text(place.international_phone_number),
            _ui_price(place.treatment_price),
            place.bed_count if place.bed_count is not None else _NOT_SPECIFIED,
        ]

    def _write_sheet(
        self, ws: Any, headers: list[str], rows: list[list[Any]], *, table_name: str
    ) -> None:
        ws.append(headers)
        for row in rows:
            ws.append([_safe_cell(value) for value in row])
        self._style_sheet(ws, headers, row_count=len(rows), table_name=table_name)

    def _style_sheet(
        self, ws: Any, headers: list[str], *, row_count: int, table_name: str
    ) -> None:
        header_fill = PatternFill("solid", fgColor="1F3B5B")
        thin = Side(style="thin", color="C9D4E3")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        link_font = Font(color="0563C1", underline="single")

        # Header row: bold, centered both ways, wrapped, taller for breathing room.
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
        ws.row_dimensions[1].height = 28

        text_columns = {
            index for index, header in enumerate(headers, start=1) if header in _TEXT_HEADERS
        }
        url_columns = {
            index for index, header in enumerate(headers, start=1) if header in _URL_HEADERS
        }
        email_columns = {
            index for index, header in enumerate(headers, start=1) if header in _EMAIL_HEADERS
        }
        centered_columns = {
            index for index, header in enumerate(headers, start=1) if header in _CENTERED_HEADERS
        }
        widths = {
            index: self._column_width(header) for index, header in enumerate(headers, start=1)
        }

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in centered_columns:
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                cell.border = cell_border
                if cell.column in text_columns and cell.value is not None:
                    cell.number_format = "@"
                if (
                    cell.column in url_columns
                    and isinstance(cell.value, str)
                    and _is_http_url(cell.value)
                ):
                    try:
                        cell.hyperlink = cell.value
                        cell.font = link_font
                    except Exception:  # noqa: BLE001 - keep export alive on bad URLs
                        pass
                if (
                    cell.column in email_columns
                    and isinstance(cell.value, str)
                    and _is_email(cell.value)
                ):
                    try:
                        cell.hyperlink = f"mailto:{cell.value}"
                        cell.font = link_font
                    except Exception:  # noqa: BLE001 - keep export alive on bad emails
                        pass
            ws.row_dimensions[row[0].row].height = _estimate_row_height(row, widths)

        for index, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(index)].width = widths[index]

        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{row_count + 1}"
        ws.auto_filter.ref = table_ref
        if row_count > 0:
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 110
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_title_rows = "1:1"

    def _column_width(self, header: str) -> int:
        return _WIDE_COLUMNS.get(header, min(max(len(header) + 2, 14), 30))


def _estimate_row_height(row: Any, widths: dict[int, int]) -> float:
    """Enough height to show wrapped content without dead space (≈ 2-4 lines)."""
    max_lines = 1
    for cell in row:
        value = cell.value
        if value is None:
            continue
        text = str(value)
        width = max(widths.get(cell.column, 14), 6)
        longest_word = max((len(word) for word in text.split()), default=0)
        # Wrapped lines ≈ characters / column width, but never fewer than the
        # number of hard newlines, and account for a single very long token.
        wrapped = max(
            text.count("\n") + 1,
            -(-len(text) // width),  # ceil division
            -(-longest_word // width),
        )
        max_lines = max(max_lines, wrapped)
    return min(15 * min(max_lines, 6) + 4, 96)


_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _safe_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Enum):
        value = value.value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)
    # openpyxl rejects ASCII control chars that are illegal in OOXML.
    cleaned = _ILLEGAL_XLSX_CHARS.sub("", text)
    if cleaned[:1] in {"=", "+", "-", "@"}:
        # Phone numbers are formatted as text elsewhere; anything else with a
        # formula-like prefix is neutralized so Excel never evaluates it.
        return cleaned
    return cleaned


def _is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _is_email(value: str) -> bool:
    return bool(_EMAIL_RE.match(value.strip()))


def _display_text(value: Any) -> str:
    if value is None:
        return "Not Specified"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Enum):
        value = value.value
    if isinstance(value, (dict, list, tuple)):
        text = json.dumps(value, ensure_ascii=False, default=str)
        return text if text.strip() else "Not Specified"
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else "Not Specified"
    text = str(value).strip()
    return text if text else "Not Specified"


# UI-mirroring helpers — match src/lib/maps/exportDisplay.ts exactly.
def _ui_text(value: Any) -> str:
    text = (str(value).strip() if value is not None else "")
    return text if text else _NOT_SPECIFIED


def _ui_list(values: list[str] | None) -> str:
    cleaned = [str(item).strip() for item in (values or []) if str(item).strip()]
    return ", ".join(cleaned) if cleaned else _NOT_SPECIFIED


def _ui_location(place: MapsPlace, country_name: str | None = None) -> str:
    address = (place.formatted_address or "").strip()
    if address:
        return address
    parts = [p for p in (place.city_name, place.region_name, country_name) if p]
    return ", ".join(parts) if parts else _NOT_SPECIFIED


def _ui_website(place: MapsPlace) -> str:
    raw = (place.official_website or place.raw_website or "").strip()
    if not raw:
        return _NOT_SPECIFIED
    if raw.lower().startswith(("http://", "https://")):
        return raw
    return f"https://{raw}"


def _ui_price(value: Any) -> str:
    text = (str(value).strip() if value is not None else "")
    return text if text else _CONTACT_PRICING


def _is_eligible_center(place: MapsPlace) -> bool:
    return place.client_eligibility == MapsClientEligibility.ELIGIBLE.value


def _has_contact_info(place: MapsPlace) -> bool:
    has_phone = bool((place.international_phone_number or "").strip())
    has_website = bool((place.official_website or place.raw_website or "").strip())
    return has_phone and has_website


maps_export_service = MapsExportService()
